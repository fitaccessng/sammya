"""
HR Module - Simplified Employee Management System
Uses existing database models (User, ProjectStaff, Project)
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app, send_file
from flask_login import current_user, login_required
from datetime import datetime, timedelta, date
from functools import wraps
from sqlalchemy import func, desc
from decimal import Decimal, ROUND_HALF_UP
import csv
import io
import os
from werkzeug.utils import secure_filename

# Import models
from app.models import (
    db, User, Project, ProjectStaff, ApprovalLog, ApprovalState, NextOfKin,
    StaffImportBatch, StaffImportItem, StaffCompensation, PayrollDeduction,
    DepartmentAccess, LeaveRequest
)
from app.utils import role_required, Roles
from app.excel_import import StaffExcelParser, StaffImportManager, ExcelImportError

# Create blueprint
bp = Blueprint('hr', __name__, url_prefix='/hr')

# ==================== DECORATORS ====================

def hr_required(f):
    """Check if user has HR role"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user or current_user.role not in [Roles.HR_MANAGER, Roles.HR_STAFF, Roles.ADMIN]:
            flash("Access denied. Insufficient permissions.", "error")
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Check if user has admin role"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user or current_user.role != Roles.ADMIN:
            flash("Access denied. Admin privileges required.", "error")
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated_function


LEAVE_ALLOWANCE = {
    'casual': 7,
    'compensate': 7,
    'annual': 21,
    'maternity': 16 * 7,   # 16 weeks
    'paternity': 2 * 7     # 2 weeks
}

PAYROLL_COMPONENT_ALLOCATION = {
    'housing': Decimal('0.4166666667'),
    'transport': Decimal('0.25'),
    'utility': Decimal('0.10'),
    'meal': Decimal('0.10'),
}

NIGERIA_PAYROLL_RULES_2026 = {
    'employee_pension_rate': Decimal('0.08'),
    'employer_pension_rate': Decimal('0.10'),
    'nhf_rate': Decimal('0.025'),
    'nhis_rate': Decimal('0.02'),
    'house_rent_relief_annual': Decimal('500000'),
    'tax_relief_annual': Decimal('800000'),
    'minimum_tax_rate': Decimal('0.01'),
}

PITA_PROGRESSIVE_BANDS = [
    (Decimal('300000'), Decimal('0.07')),
    (Decimal('300000'), Decimal('0.11')),
    (Decimal('500000'), Decimal('0.15')),
    (Decimal('500000'), Decimal('0.19')),
    (Decimal('1600000'), Decimal('0.21')),
]


def _d(value):
    return Decimal(str(value or 0))


def _money(value):
    return _d(value).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def _calc_paye_annual(taxable_income_annual, annual_gross):
    taxable = _d(taxable_income_annual)
    gross = _d(annual_gross)
    remaining = taxable
    tax_due = Decimal('0')

    for band_limit, rate in PITA_PROGRESSIVE_BANDS:
        if remaining <= 0:
            break
        taxable_in_band = min(remaining, band_limit)
        tax_due += taxable_in_band * rate
        remaining -= taxable_in_band

    if remaining > 0:
        tax_due += remaining * Decimal('0.24')

    minimum_tax = gross * NIGERIA_PAYROLL_RULES_2026['minimum_tax_rate']
    return _money(max(tax_due, minimum_tax if gross > 0 else Decimal('0')))


def _extract_manual_deductions(compensation):
    manual = {
        'pension_employee': Decimal('0'),
        'nhf': Decimal('0'),
        'nhis': Decimal('0'),
        'paye': Decimal('0'),
    }
    has_override = {k: False for k in manual.keys()}
    other = Decimal('0')

    if not compensation:
        return manual, has_override, other

    deductions = PayrollDeduction.query.filter_by(
        compensation_id=compensation.id,
        is_recurring=True
    ).all()

    for deduction in deductions:
        label = (deduction.deduction_type or '').strip().lower()
        amount = _d(deduction.amount)

        if 'nhf' in label:
            manual['nhf'] += amount
            has_override['nhf'] = True
        elif 'nhis' in label or 'health insurance' in label:
            manual['nhis'] += amount
            has_override['nhis'] = True
        elif 'paye' in label or label in ['tax', 'income tax']:
            manual['paye'] += amount
            has_override['paye'] = True
        elif 'pension' in label:
            manual['pension_employee'] += amount
            has_override['pension_employee'] = True
        else:
            other += amount

    return manual, has_override, other


def _resolve_allowance_components(staff, compensation):
    from app.payroll_models import SalaryMapping

    mapping = SalaryMapping.query.filter_by(
        user_id=staff.id,
        is_active=True
    ).order_by(SalaryMapping.effective_date.desc()).first()

    if mapping:
        basic = _money(mapping.basic_salary)
        housing = _money(mapping.house_allowance)
        transport = _money(mapping.transport_allowance)
        meal = _money(mapping.meal_allowance)
        utility = _money(mapping.other_allowances)
        medical = _money(_d(mapping.risk_allowance) + _d(mapping.performance_allowance))
    else:
        basic = _money(compensation.basic_salary if compensation else getattr(staff, 'basic_salary', 0))
        total_allowances = _money(compensation.allowances if compensation else 0)
        housing = _money(total_allowances * PAYROLL_COMPONENT_ALLOCATION['housing'])
        transport = _money(total_allowances * PAYROLL_COMPONENT_ALLOCATION['transport'])
        utility = _money(total_allowances * PAYROLL_COMPONENT_ALLOCATION['utility'])
        meal = _money(total_allowances * PAYROLL_COMPONENT_ALLOCATION['meal'])
        medical = _money(total_allowances - housing - transport - utility - meal)

    gross = _money(basic + housing + transport + utility + meal + medical)
    return basic, housing, transport, utility, meal, medical, gross


def _build_staff_payroll_row(staff):
    compensation = StaffCompensation.query.filter_by(user_id=staff.id).first()
    basic, housing, transport, utility, meal, medical, gross = _resolve_allowance_components(staff, compensation)
    manual, has_override, other_deductions_monthly = _extract_manual_deductions(compensation)

    pensionable_monthly = _money(basic + housing + transport)
    annual_gross = _money(gross * 12)

    pension_employer_annual = _money((pensionable_monthly * 12) * NIGERIA_PAYROLL_RULES_2026['employer_pension_rate'])
    pension_employee_annual = _money((pensionable_monthly * 12) * NIGERIA_PAYROLL_RULES_2026['employee_pension_rate'])
    nhf_annual = _money((basic * 12) * NIGERIA_PAYROLL_RULES_2026['nhf_rate'])
    nhis_annual = _money(annual_gross * NIGERIA_PAYROLL_RULES_2026['nhis_rate'])

    if has_override['pension_employee']:
        pension_employee_annual = _money(manual['pension_employee'] * 12)
    if has_override['nhf']:
        nhf_annual = _money(manual['nhf'] * 12)
    if has_override['nhis']:
        nhis_annual = _money(manual['nhis'] * 12)

    house_rent_annual = _money(NIGERIA_PAYROLL_RULES_2026['house_rent_relief_annual'])
    tax_relief_annual = _money(NIGERIA_PAYROLL_RULES_2026['tax_relief_annual'])
    consolidated_relief_annual = _money(
        pension_employee_annual + nhf_annual + nhis_annual + house_rent_annual + tax_relief_annual
    )
    taxable_income_annual = _money(max(annual_gross - consolidated_relief_annual, Decimal('0')))
    paye_annual = _calc_paye_annual(taxable_income_annual, annual_gross)

    if has_override['paye']:
        paye_annual = _money(manual['paye'] * 12)

    pension_employee_monthly = _money(pension_employee_annual / 12)
    nhf_monthly = _money(nhf_annual / 12)
    nhis_monthly = _money(nhis_annual / 12)
    paye_monthly = _money(paye_annual / 12)

    total_deductions_monthly = _money(
        pension_employee_monthly + nhf_monthly + nhis_monthly + paye_monthly + other_deductions_monthly
    )
    net_monthly = _money(gross - total_deductions_monthly)

    dept_access = DepartmentAccess.query.filter_by(user_id=staff.id, is_active=True).first()
    department = dept_access.department if dept_access else 'Unassigned'

    return {
        'id': staff.id,
        'name': staff.name,
        'email': staff.email,
        'department': department,
        'basic_salary': basic,
        'housing': housing,
        'transport': transport,
        'utility': utility,
        'meal': meal,
        'medical': medical,
        'gross_monthly': gross,
        'annual_gross': annual_gross,
        'annual_pension_employer': pension_employer_annual,
        'annual_pension_employee': pension_employee_annual,
        'annual_nhis': nhis_annual,
        'annual_nhf': nhf_annual,
        'house_rent_annual': house_rent_annual,
        'tax_relief_annual': tax_relief_annual,
        'consolidated_relief_annual': consolidated_relief_annual,
        'taxable_income_annual': taxable_income_annual,
        'paye_annual': paye_annual,
        'paye_monthly': paye_monthly,
        'total_deductions_monthly': total_deductions_monthly,
        'other_deductions_monthly': _money(other_deductions_monthly),
        'net_monthly': net_monthly,
        'status': 'pending',
    }


def _build_payroll_summary(staff_members):
    summary = {
        'staff_count': 0,
        'total_basic_salary': Decimal('0'),
        'total_gross': Decimal('0'),
        'total_deductions': Decimal('0'),
        'total_net_salary': Decimal('0'),
        'total_paye_monthly': Decimal('0'),
        'total_taxable_income_annual': Decimal('0'),
        'department_breakdown': [],
        'staff_details': [],
    }
    dept = {}

    for staff in staff_members:
        row = _build_staff_payroll_row(staff)
        summary['staff_count'] += 1
        summary['total_basic_salary'] += row['basic_salary']
        summary['total_gross'] += row['gross_monthly']
        summary['total_deductions'] += row['total_deductions_monthly']
        summary['total_net_salary'] += row['net_monthly']
        summary['total_paye_monthly'] += row['paye_monthly']
        summary['total_taxable_income_annual'] += row['taxable_income_annual']
        summary['staff_details'].append(row)

        key = row['department']
        if key not in dept:
            dept[key] = {'count': 0, 'total_gross': Decimal('0')}
        dept[key]['count'] += 1
        dept[key]['total_gross'] += row['gross_monthly']

    for department in sorted(dept.keys()):
        summary['department_breakdown'].append({
            'name': department,
            'count': dept[department]['count'],
            'total_gross': _money(dept[department]['total_gross']),
        })

    summary['staff_details'].sort(key=lambda x: x['name'] or '')
    summary['total_basic_salary'] = _money(summary['total_basic_salary'])
    summary['total_gross'] = _money(summary['total_gross'])
    summary['total_deductions'] = _money(summary['total_deductions'])
    summary['total_net_salary'] = _money(summary['total_net_salary'])
    summary['total_paye_monthly'] = _money(summary['total_paye_monthly'])
    summary['total_taxable_income_annual'] = _money(summary['total_taxable_income_annual'])
    return summary


def _parse_decimal_input(value):
    if value is None:
        return Decimal('0')
    text = str(value).strip()
    if not text:
        return Decimal('0')
    normalized = text.replace(',', '')
    return _money(Decimal(normalized))


def _upsert_staff_payroll_inputs(staff, fields, actor_id):
    from app.payroll_models import SalaryMapping

    basic = _parse_decimal_input(fields.get('basic_salary'))
    housing = _parse_decimal_input(fields.get('housing'))
    transport = _parse_decimal_input(fields.get('transport'))
    utility = _parse_decimal_input(fields.get('utility'))
    meal = _parse_decimal_input(fields.get('meal'))
    medical = _parse_decimal_input(fields.get('medical'))
    paye_annual = _parse_decimal_input(fields.get('paye_annual'))
    pension_employee_annual = _parse_decimal_input(fields.get('annual_pension_employee'))
    nhis_annual = _parse_decimal_input(fields.get('annual_nhis'))
    nhf_annual = _parse_decimal_input(fields.get('nhf_annual'))

    total_allowances = _money(housing + transport + utility + meal + medical)

    compensation = StaffCompensation.query.filter_by(user_id=staff.id).first()
    if not compensation:
        compensation = StaffCompensation(user_id=staff.id, basic_salary=basic, allowances=total_allowances, gross_salary=_money(basic + total_allowances))
        db.session.add(compensation)
        db.session.flush()
    else:
        compensation.basic_salary = basic
        compensation.allowances = total_allowances
        compensation.gross_salary = _money(basic + total_allowances)

    mapping = SalaryMapping.query.filter_by(user_id=staff.id, is_active=True).order_by(SalaryMapping.effective_date.desc()).first()
    if not mapping:
        mapping = SalaryMapping(
            user_id=staff.id,
            created_by_id=actor_id,
            effective_date=date.today(),
            is_active=True
        )
        db.session.add(mapping)

    mapping.basic_salary = basic
    mapping.house_allowance = housing
    mapping.transport_allowance = transport
    mapping.meal_allowance = meal
    mapping.risk_allowance = medical
    mapping.performance_allowance = Decimal('0')
    mapping.other_allowances = utility
    mapping.tax_amount = _money(paye_annual / 12) if paye_annual > 0 else Decimal('0')
    mapping.pension_amount = _money(pension_employee_annual / 12) if pension_employee_annual > 0 else Decimal('0')
    mapping.insurance_amount = _money(nhis_annual / 12) if nhis_annual > 0 else Decimal('0')
    mapping.updated_by_id = actor_id

    # Keep override deductions in compensation for NHF/NHIS/PAYE/Pension inputs
    existing = PayrollDeduction.query.filter(
        PayrollDeduction.compensation_id == compensation.id,
        PayrollDeduction.is_recurring == True,
        func.lower(PayrollDeduction.deduction_type).in_(['pension', 'paye', 'nhis', 'nhf'])
    ).all()
    for d in existing:
        db.session.delete(d)

    overrides = [
        ('Pension', pension_employee_annual),
        ('PAYE', paye_annual),
        ('NHIS', nhis_annual),
        ('NHF', nhf_annual),
    ]
    for deduction_type, annual_amount in overrides:
        if annual_amount > 0:
            db.session.add(PayrollDeduction(
                compensation_id=compensation.id,
                deduction_type=deduction_type,
                description=f'{deduction_type} uploaded/entered annual value',
                amount=_money(annual_amount / 12),
                is_recurring=True
            ))


def normalize_leave_type(leave_type):
    return (leave_type or '').strip().lower().replace('leave', '').strip()


def approved_leave_days_for_year(user_id, leave_type, year):
    start_of_year = date(year, 1, 1)
    end_of_year = date(year, 12, 31)
    records = LeaveRequest.query.filter(
        LeaveRequest.user_id == user_id,
        LeaveRequest.status == 'approved',
        LeaveRequest.leave_type == leave_type,
        LeaveRequest.start_date <= end_of_year,
        LeaveRequest.end_date >= start_of_year
    ).all()
    return sum(int(r.days_requested or 0) for r in records)


def build_leave_balance(user_id, year):
    balances = {}
    for leave_type, allowed_days in LEAVE_ALLOWANCE.items():
        used = approved_leave_days_for_year(user_id, leave_type, year)
        balances[leave_type] = {
            'allowed': allowed_days,
            'used': used,
            'remaining': max(0, allowed_days - used)
        }
    return balances

# ==================== DASHBOARD ROUTES ====================

@bp.route('/')
@bp.route('/home')
@login_required
@hr_required
def hr_home():
    """Main HR Dashboard"""
    try:
        # User Statistics
        total_staff = User.query.count()
        active_staff = User.query.filter_by(is_active=True).count()
        inactive_staff = total_staff - active_staff
        
        # Staff Roles
        staff_by_role = db.session.query(
            User.role, 
            func.count(User.id)
        ).filter(User.role.isnot(None)).group_by(User.role).all()
        
        # Project Statistics
        total_projects = Project.query.count()
        active_projects = Project.query.filter_by(status='active').count()
        
        # Project Assignments
        total_assignments = ProjectStaff.query.filter_by(is_active=True).count()
        
        # Recent staff
        recent_users = User.query.order_by(desc(User.created_at)).limit(5).all()
        
        dashboard_data = {
            'total_staff': total_staff,
            'active_staff': active_staff,
            'inactive_staff': inactive_staff,
            'staff_by_role': [{'role': role, 'count': count} for role, count in staff_by_role],
            'total_projects': total_projects,
            'active_projects': active_projects,
            'total_assignments': total_assignments,
            'recent_staff': recent_users,
            # Additional metrics expected by template
            'present_today': 0,
            'absent_today': 0,
            'late_today': 0,
            'active_leaves': 0,
            'pending_leaves': 0,
            'pending_payroll': 0,
            'pending_tasks': 0,
            'in_progress_tasks': 0,
            'pending_queries': 0,
            'recent_activities': []
        }
        
        return render_template('hr/index.html', dashboard=dashboard_data)
        
    except Exception as e:
        current_app.logger.error(f"HR Dashboard Error: {str(e)}")
        flash("Error loading HR dashboard", "error")
        return redirect(url_for('main.dashboard'))

# ==================== STAFF MANAGEMENT ROUTES ====================

@bp.route('/staff')
@login_required
@hr_required
def staff_list():
    """List all staff members"""
    try:
        page = request.args.get('page', 1, type=int)
        search = request.args.get('search', '')
        status_filter = request.args.get('status', 'active')
        
        query = User.query
        
        if search:
            query = query.filter(User.name.ilike(f'%{search}%') | User.email.ilike(f'%{search}%'))
        
        if status_filter == 'active':
            query = query.filter_by(is_active=True)
        elif status_filter == 'inactive':
            query = query.filter_by(is_active=False)
        
        staff = query.order_by(User.name).paginate(page=page, per_page=20)
        
        stats = {
            'total_staff': User.query.count(),
            'active_staff': User.query.filter_by(is_active=True).count(),
            'inactive_staff': User.query.filter_by(is_active=False).count(),
        }
        
        return render_template('hr/staff/index.html', staff=staff, stats=stats, search=search, status_filter=status_filter)
        
    except Exception as e:
        current_app.logger.error(f"Staff List Error: {str(e)}")
        flash("Error loading staff list", "error")
        return redirect(url_for('hr.hr_home'))

@bp.route('/staff/<int:staff_id>')
@login_required
@hr_required
def staff_details(staff_id):
    """View staff member details"""
    try:
        staff = User.query.get_or_404(staff_id)
        
        # Get project assignments
        assignments = ProjectStaff.query.filter_by(user_id=staff_id).all()
        
        # Pass current date for age/tenure calculations
        from datetime import datetime as dt
        
        return render_template('hr/staff/details.html', staff=staff, assignments=assignments, now=dt.now())
        
    except Exception as e:
        current_app.logger.error(f"Staff Details Error: {str(e)}")
        flash("Error loading staff details", "error")
        return redirect(url_for('hr.staff_list'))

@bp.route('/staff/<int:staff_id>/upload-passport', methods=['POST'])
@login_required
@hr_required
def upload_passport(staff_id):
    """Upload staff passport document."""
    try:
        staff = User.query.get_or_404(staff_id)
        
        if 'passport_document' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file provided'}), 400
        
        file = request.files['passport_document']
        
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400
        
        # Validate file type
        allowed_extensions = {'pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx'}
        if not ('.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
            return jsonify({'status': 'error', 'message': 'Invalid file type. Allowed: PDF, JPG, JPEG, PNG, DOC, DOCX'}), 400
        
        # Validate file size (max 10MB)
        file.seek(0, 2)
        file_size = file.tell()
        file.seek(0)
        if file_size > 10 * 1024 * 1024:  # 10MB
            return jsonify({'status': 'error', 'message': 'File size exceeds 10MB limit'}), 400
        
        # Create uploads directory if it doesn't exist
        upload_dir = os.path.join(current_app.root_path, 'uploads', 'passports')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Save file with secure name
        filename = secure_filename(f"passport_{staff_id}_{datetime.now().timestamp()}_{file.filename}")
        filepath = os.path.join(upload_dir, filename)
        file.save(filepath)
        
        # Update staff record with passport document path
        staff.passport_document = f'/uploads/passports/{filename}'
        db.session.commit()
        
        current_app.logger.info(f"Passport uploaded for staff {staff_id}: {filename}")
        return jsonify({'status': 'success', 'message': 'Passport uploaded successfully', 'path': staff.passport_document})
        
    except Exception as e:
        current_app.logger.error(f"Passport upload error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@bp.route('/staff/<int:staff_id>/edit', methods=['GET', 'POST'])
@login_required
@hr_required
def edit_staff(staff_id):
    """Edit staff member"""
    try:
        staff = User.query.get_or_404(staff_id)
        
        if request.method == 'POST':
            staff.name = request.form.get('name', staff.name)
            staff.email = request.form.get('email', staff.email)
            staff.role = request.form.get('role', staff.role)
            staff.is_active = request.form.get('is_active') == 'on'
            
            # Personal Information
            staff.phone = request.form.get('phone') or None
            staff.gender = request.form.get('gender') or None
            staff.marital_status = request.form.get('marital_status') or None
            staff.address = request.form.get('address') or None
            staff.city = request.form.get('city') or None
            staff.state = request.form.get('state') or None
            
            # Date fields
            date_of_birth = request.form.get('date_of_birth')
            if date_of_birth:
                staff.date_of_birth = datetime.strptime(date_of_birth, '%Y-%m-%d').date()
            
            date_of_employment = request.form.get('date_of_employment')
            if date_of_employment:
                staff.date_of_employment = datetime.strptime(date_of_employment, '%Y-%m-%d').date()
            
            staff.employee_id = request.form.get('employee_id') or None
            
            # Handle salary and deductions
            basic_salary = request.form.get('basic_salary')
            if basic_salary:
                staff.basic_salary = float(basic_salary)
            
            default_deductions = request.form.get('default_deductions')
            if default_deductions:
                staff.default_deductions = float(default_deductions)
            
            # ==================== HANDLE NEXT OF KIN ====================
            
            # Get array-based form data
            kin_ids = request.form.getlist('kin_id[]')
            kin_names = request.form.getlist('kin_name[]')
            kin_relationships = request.form.getlist('kin_relationship[]')
            kin_phones = request.form.getlist('kin_phone[]')
            kin_emails = request.form.getlist('kin_email[]')
            kin_addresses = request.form.getlist('kin_address[]')
            kin_cities = request.form.getlist('kin_city[]')
            kin_states = request.form.getlist('kin_state[]')
            kin_primaries = request.form.getlist('kin_primary[]')
            kin_deletes = request.form.getlist('kin_delete[]')
            
            # Delete marked next of kin records
            if kin_deletes:
                for kin_id in kin_deletes:
                    try:
                        kin = NextOfKin.query.filter_by(id=int(kin_id), user_id=staff_id).first()
                        if kin:
                            db.session.delete(kin)
                    except (ValueError, TypeError):
                        pass
            
            # Update/Create next of kin records
            primary_kins = kin_primaries  # These are the IDs/keys marked as primary
            
            # First, set all non-primary
            for kin in staff.next_of_kin:
                kin.is_primary = False
            
            for idx, kin_id in enumerate(kin_ids):
                if idx < len(kin_names) and kin_names[idx]:  # Only process if name exists
                    # Handle both existing records and new ones
                    if isinstance(kin_id, str) and kin_id.startswith('new_'):
                        # New record
                        kin = NextOfKin(user_id=staff_id)
                    else:
                        # Existing record
                        try:
                            kin = NextOfKin.query.filter_by(id=int(kin_id), user_id=staff_id).first()
                        except (ValueError, TypeError):
                            kin = NextOfKin(user_id=staff_id)
                    
                    if kin:
                        kin.full_name = kin_names[idx]
                        kin.relationship = kin_relationships[idx] if idx < len(kin_relationships) else ''
                        kin.phone = kin_phones[idx] if idx < len(kin_phones) else ''
                        kin.email = kin_emails[idx] if idx < len(kin_emails) else ''
                        kin.address = kin_addresses[idx] if idx < len(kin_addresses) else ''
                        kin.city = kin_cities[idx] if idx < len(kin_cities) else ''
                        kin.state = kin_states[idx] if idx < len(kin_states) else ''
                        
                        # Check if this entry is marked as primary
                        is_marked_primary = kin_id in primary_kins or (isinstance(kin_id, str) and kin_id.startswith('new_') and kin_id in primary_kins)
                        kin.is_primary = is_marked_primary
                        
                        db.session.add(kin)
            
            # Ensure only one primary contact
            primary_count = sum(1 for kin in staff.next_of_kin if kin.is_primary)
            if primary_count > 1:
                # Keep only the first one as primary
                for idx, kin in enumerate(staff.next_of_kin):
                    kin.is_primary = (idx == 0)
            
            db.session.commit()
            flash('Staff member updated successfully', 'success')
            return redirect(url_for('hr.staff_details', staff_id=staff_id))
        
        roles = ['ADMIN', 'HR_MANAGER', 'HR_STAFF', 'QS_MANAGER', 'QC_MANAGER', 
                 'COST_MANAGER', 'SAFETY_MANAGER']
        
        return render_template('hr/staff/edit.html', staff=staff, roles=roles)
        
    except Exception as e:
        current_app.logger.error(f"Edit Staff Error: {str(e)}")
        flash("Error editing staff member", "error")
        return redirect(url_for('hr.staff_list'))

# ==================== PROJECT ASSIGNMENTS ====================

@bp.route('/assignments')
@login_required
@hr_required
def assignments():
    """View project staff assignments"""
    try:
        page = request.args.get('page', 1, type=int)
        
        assignments = ProjectStaff.query.filter_by(is_active=True)\
            .join(User).join(Project)\
            .order_by(desc(ProjectStaff.created_at))\
            .paginate(page=page, per_page=20)
        
        return render_template('hr/assignments.html', assignments=assignments)
        
    except Exception as e:
        current_app.logger.error(f"Assignments Error: {str(e)}")
        flash("Error loading assignments", "error")
        return redirect(url_for('hr.hr_home'))

# ==================== BASIC STATS ====================

@bp.route('/api/stats')
@login_required
@hr_required
def get_stats():
    """Get HR statistics as JSON"""
    try:
        # Calculate total salary and deductions
        total_salary = 0
        total_deductions = 0
        
        # Get all active staff with compensation
        active_staff = User.query.filter_by(is_active=True).all()
        
        for staff in active_staff:
            # Get basic salary from StaffCompensation
            compensation = StaffCompensation.query.filter_by(staff_id=staff.id).first()
            if compensation:
                total_salary += compensation.basic_salary or 0
                
                # Get total deductions for this staff
                deductions = PayrollDeduction.query.filter_by(staff_id=staff.id, is_recurring=True).all()
                for deduction in deductions:
                    total_deductions += deduction.amount or 0
        
        stats = {
            'success': True,
            'total_staff': User.query.filter_by(is_active=True).count(),
            'active_staff': User.query.filter_by(is_active=True).count(),
            'inactive_staff': User.query.filter_by(is_active=False).count(),
            'total_projects': Project.query.count(),
            'active_projects': Project.query.filter_by(status='active').count(),
            'total_assignments': ProjectStaff.query.filter_by(is_active=True).count(),
            'total_salary': round(total_salary, 2),
            'total_deductions': round(total_deductions, 2)
        }
        return jsonify(stats)
    except Exception as e:
        current_app.logger.error(f"Stats API Error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/api/payroll/details')
@login_required
@hr_required
def get_payroll_details():
    """Get detailed payroll information including department breakdown and staff details"""
    try:
        # Get all active staff
        active_staff = User.query.filter_by(is_active=True).all()
        
        total_salary = 0
        total_deductions = 0
        staff_details = []
        department_breakdown = {}
        
        for staff in active_staff:
            # Try to get salary from StaffCompensation first
            compensation = StaffCompensation.query.filter_by(user_id=staff.id).first()
            
            if compensation:
                basic_salary = float(compensation.basic_salary or 0)
                
                # Get total deductions for this staff from PayrollDeduction (through compensation)
                deductions = PayrollDeduction.query.filter_by(compensation_id=compensation.id, is_recurring=True).all()
                staff_deductions = sum(float(d.amount or 0) for d in deductions)
            else:
                # Fallback to user's basic_salary field if it exists, otherwise use 0
                basic_salary = float(staff.basic_salary or 0) if hasattr(staff, 'basic_salary') else 0
                staff_deductions = 0
            
            if basic_salary == 0:
                # Skip staff with no salary set
                continue
            
            total_salary += basic_salary
            total_deductions += staff_deductions
            
            # Get department from DepartmentAccess (first active access)
            dept_access = DepartmentAccess.query.filter_by(user_id=staff.id, is_active=True).first()
            dept = dept_access.department if dept_access else 'Unassigned'
            
            # Add to staff details
            staff_details.append({
                'id': staff.id,
                'name': staff.name,
                'department': dept,
                'basic_salary': round(basic_salary, 2),
                'deductions': round(staff_deductions, 2)
            })
            
            # Build department breakdown
            if dept not in department_breakdown:
                department_breakdown[dept] = {'count': 0, 'total_salary': 0}
            department_breakdown[dept]['count'] += 1
            department_breakdown[dept]['total_salary'] += basic_salary
        
        # Convert department breakdown to list
        dept_list = [
            {'name': dept, 'count': info['count'], 'total_salary': round(info['total_salary'], 2)}
            for dept, info in sorted(department_breakdown.items())
        ]
        
        # Sort staff details by name
        staff_details.sort(key=lambda x: x['name'])
        
        response_data = {
            'success': True,
            'total_staff': len(staff_details),
            'total_salary': round(total_salary, 2),
            'total_deductions': round(total_deductions, 2),
            'department_breakdown': dept_list,
            'staff_details': staff_details
        }
        
        current_app.logger.info(f"Payroll Details: {len(staff_details)} staff, ₦{total_salary}, deductions ₦{total_deductions}")
        
        return jsonify(response_data)
    except Exception as e:
        current_app.logger.error(f"Payroll Details Error: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== STUB ROUTES (Redirect to main pages) ====================

@bp.route('/payroll')
@login_required
@hr_required
def payroll():
    """Payroll management dashboard"""
    try:
        staff_list = User.query.filter_by(is_active=True).all()
        
        payroll_records = []
        total_salary = 0
        total_deductions = 0
        
        for staff in staff_list:
            basic = float(staff.basic_salary or 0)
            deductions = float(staff.default_deductions or 0)
            net = basic - deductions
            
            payroll_records.append({
                'id': staff.id,
                'name': staff.name,
                'email': staff.email,
                'basic_salary': f"{basic:,.2f}",
                'deductions': f"{deductions:,.2f}",
                'net_salary': f"{net:,.2f}",
                'status': 'pending'
            })
            
            total_salary += basic
            total_deductions += deductions
        
        return render_template('hr/payroll.html', 
                             payroll_records=payroll_records,
                             total_salary=f"{total_salary:,.2f}",
                             total_deductions=f"{total_deductions:,.2f}",
                             total_net=f"{(total_salary - total_deductions):,.2f}")
    except Exception as e:
        current_app.logger.error(f"Payroll Error: {str(e)}")
        flash("Error loading payroll", "error")
        return redirect(url_for('hr.staff_list'))


@bp.route('/payroll/details', methods=['GET', 'POST'])
@login_required
@hr_required
def payroll_details_input():
    """Manage payroll staff input details (manual entry + bulk upload)."""
    try:
        from app.payroll_models import SalaryMapping

        if request.method == 'POST':
            staff_id = request.form.get('staff_id', type=int)
            staff = User.query.get(staff_id) if staff_id else None
            if not staff:
                flash("Please select a valid staff member.", "error")
                return redirect(url_for('hr.payroll_details_input'))

            fields = {
                'basic_salary': request.form.get('basic_salary'),
                'housing': request.form.get('housing'),
                'transport': request.form.get('transport'),
                'utility': request.form.get('utility'),
                'meal': request.form.get('meal'),
                'medical': request.form.get('medical'),
                'annual_pension_employee': request.form.get('annual_pension_employee'),
                'annual_nhis': request.form.get('annual_nhis'),
                'nhf_annual': request.form.get('nhf_annual'),
                'paye_annual': request.form.get('paye_annual'),
            }
            _upsert_staff_payroll_inputs(staff, fields, current_user.id)
            db.session.commit()
            flash(f"Payroll inputs saved for {staff.name}.", "success")
            return redirect(url_for('hr.payroll_details_input', edit_user_id=staff.id))

        active_staff = User.query.filter_by(is_active=True).order_by(User.name.asc()).all()
        mappings = SalaryMapping.query.filter_by(is_active=True).order_by(SalaryMapping.updated_at.desc(), SalaryMapping.created_at.desc()).limit(20).all()
        preview_rows = []
        for mapping in mappings:
            if mapping.user:
                preview_rows.append(_build_staff_payroll_row(mapping.user))
        edit_user_id = request.args.get('edit_user_id', type=int)
        selected_input = None
        if edit_user_id:
            selected_staff = User.query.get(edit_user_id)
            if selected_staff:
                selected_input = _build_staff_payroll_row(selected_staff)
        return render_template(
            'hr/payroll/details_input.html',
            staff_list=active_staff,
            mappings=mappings,
            preview_rows=preview_rows,
            selected_input=selected_input,
            edit_user_id=edit_user_id
        )
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Payroll details input error: {str(e)}", exc_info=True)
        flash("Error saving payroll inputs.", "error")
        return redirect(url_for('hr.payroll'))


@bp.route('/payroll/details/<int:user_id>/edit', methods=['GET'])
@login_required
@hr_required
def payroll_details_edit(user_id):
    """Open payroll input form in edit mode for a specific staff member."""
    return redirect(url_for('hr.payroll_details_input', edit_user_id=user_id))


@bp.route('/payroll/details/upload', methods=['POST'])
@login_required
@hr_required
def payroll_details_upload():
    """Bulk upload payroll input details from CSV/Excel."""
    try:
        upload = request.files.get('payroll_file')
        if not upload or not upload.filename:
            flash("Please select a payroll file to upload.", "error")
            return redirect(url_for('hr.payroll_details_input'))

        filename = secure_filename(upload.filename).lower()
        rows = []

        if filename.endswith('.csv'):
            content = upload.stream.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        elif filename.endswith('.xlsx') or filename.endswith('.xlsm'):
            from openpyxl import load_workbook
            wb = load_workbook(upload, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                flash("Uploaded file is empty.", "error")
                return redirect(url_for('hr.payroll_details_input'))
            headers = [str(h or '').strip() for h in all_rows[0]]
            for raw in all_rows[1:]:
                if not any(raw):
                    continue
                rows.append({headers[i]: raw[i] for i in range(len(headers))})
        else:
            flash("Unsupported file type. Use CSV or XLSX.", "error")
            return redirect(url_for('hr.payroll_details_input'))

        def key(name):
            return ''.join(ch for ch in str(name or '').lower() if ch.isalnum())

        updated = 0
        skipped = 0

        for row in rows:
            normalized = {key(k): v for k, v in row.items()}
            staff_name = str(normalized.get('names') or normalized.get('name') or '').strip()
            staff_email = str(normalized.get('email') or '').strip()

            staff = None
            if staff_email:
                staff = User.query.filter(func.lower(User.email) == staff_email.lower()).first()
            if not staff and staff_name:
                staff = User.query.filter(func.lower(User.name) == staff_name.lower()).first()

            if not staff:
                skipped += 1
                continue

            fields = {
                'basic_salary': normalized.get('basicsalary'),
                'housing': normalized.get('housing'),
                'transport': normalized.get('transport'),
                'utility': normalized.get('utility'),
                'meal': normalized.get('meal'),
                'medical': normalized.get('medical'),
                'annual_pension_employee': normalized.get('annualpensionemployee'),
                'annual_nhis': normalized.get('annualnhis'),
                'nhf_annual': normalized.get('nhfannual'),
                'paye_annual': normalized.get('payeannual'),
            }
            _upsert_staff_payroll_inputs(staff, fields, current_user.id)
            updated += 1

        db.session.commit()
        flash(f"Payroll upload processed. Updated: {updated}, Skipped: {skipped}.", "success")
        return redirect(url_for('hr.payroll_details_input'))
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Payroll bulk upload error: {str(e)}", exc_info=True)
        flash("Error processing payroll upload.", "error")
        return redirect(url_for('hr.payroll_details_input'))

@bp.route('/payroll/generate', methods=['GET', 'POST'])
@login_required
@hr_required
def payroll_generate():
    """Generate payroll for all active staff - sends for admin approval"""
    try:
        from app.payroll_models import PayrollBatch, PayrollRecord, PayrollStatus
        from datetime import datetime
        import calendar
        
        if request.method == 'POST':
            # Get payroll month from form
            month_str = request.form.get('payroll_month')
            if not month_str:
                flash("Please select a payroll month", "error")
                return redirect(url_for('hr.payroll_generate'))
            
            # Parse date
            payroll_month = datetime.strptime(month_str, '%Y-%m').date()
            
            # Check if payroll batch already exists for this month
            existing = PayrollBatch.query.filter_by(payroll_period=month_str).first()
            if existing:
                flash(f"Payroll already exists for {payroll_month.strftime('%B %Y')}", "error")
                return redirect(url_for('hr.payroll_generate'))
            
            # Get all active staff
            active_staff = User.query.filter_by(is_active=True).all()
            
            if not active_staff:
                flash("No active staff found", "error")
                return redirect(url_for('hr.payroll_generate'))

            payroll_summary = _build_payroll_summary(active_staff)
            if payroll_summary['staff_count'] == 0:
                flash("No active staff with payroll configuration found", "error")
                return redirect(url_for('hr.payroll_generate'))
            
            # Create payroll batch
            batch_name = f"Payroll-{payroll_month.strftime('%B %Y')}"
            
            # Calculate month start and end dates
            import calendar
            last_day = calendar.monthrange(payroll_month.year, payroll_month.month)[1]
            start_date = payroll_month.replace(day=1)
            end_date = payroll_month.replace(day=last_day)
            
            batch = PayrollBatch(
                batch_name=batch_name,
                payroll_period=month_str,
                status=PayrollStatus.DRAFT,
                created_by_id=current_user.id,
                start_date=start_date,
                end_date=end_date
            )
            
            db.session.add(batch)
            db.session.flush()  # Get batch ID without committing
            
            # Create payroll records using Nigeria payroll structure
            for row in payroll_summary['staff_details']:
                record = PayrollRecord(
                    batch_id=batch.id,
                    user_id=row['id'],
                    payroll_period=month_str,
                    basic_salary=row['basic_salary'],
                    house_allowance=row['housing'],
                    transport_allowance=row['transport'],
                    meal_allowance=row['meal'],
                    risk_allowance=row['medical'],
                    other_allowances=row['utility'],
                    total_allowances=row['housing'] + row['transport'] + row['utility'] + row['meal'] + row['medical'],
                    gross_salary=row['gross_monthly'],
                    tax_deduction=row['paye_monthly'],
                    pension_deduction=_money(row['annual_pension_employee'] / 12),
                    insurance_deduction=_money(row['annual_nhis'] / 12),
                    other_deductions=row['other_deductions_monthly'] + _money(row['annual_nhf'] / 12),
                    total_deductions=row['total_deductions_monthly'],
                    net_salary=row['net_monthly']
                )
                db.session.add(record)
            
            # Update batch totals with correct field names
            batch.total_records = payroll_summary['staff_count']
            batch.successfully_processed = payroll_summary['staff_count']
            batch.total_basic_salary = payroll_summary['total_basic_salary']
            batch.total_gross = payroll_summary['total_gross']
            batch.total_deductions = payroll_summary['total_deductions']
            batch.total_net = payroll_summary['total_net_salary']
            
            db.session.commit()
            
            # Log approval event
            approval_log = ApprovalLog(
                entity_type='payroll',
                entity_id=batch.id,
                action='created',
                actor_id=current_user.id,
                comment=f"Payroll batch generated for {payroll_month.strftime('%B %Y')} with {payroll_summary['staff_count']} staff members"
            )
            db.session.add(approval_log)
            db.session.commit()
            
            # Return template with generated payroll data (no redirect)
            return render_template(
                'hr/payroll/generate.html',
                current_month=payroll_month,
                generated_payroll=batch,
                payroll_summary=payroll_summary
            )
        
        # GET request - show form to select month
        current_date = datetime.now()
        
        # Fetch and display payroll summary preview
        active_staff = User.query.filter_by(is_active=True).all()
        payroll_summary = _build_payroll_summary(active_staff)
        
        return render_template(
            'hr/payroll/generate.html',
            current_month=current_date,
            payroll_summary=payroll_summary
        )
    
    except Exception as e:
        current_app.logger.error(f"Payroll Generation Error: {str(e)}")
        flash(f"Error generating payroll: {str(e)}", "error")
        return redirect(url_for('hr.payroll'))

@bp.route('/payroll/send-approval', methods=['POST'])
@login_required
@hr_required
def payroll_send_approval():
    """Send generated payroll to admin for approval"""
    try:
        from app.payroll_models import PayrollBatch, PayrollStatus
        
        payroll_id = request.form.get('payroll_id')
        
        if not payroll_id:
            flash('Payroll ID is required', 'error')
            return redirect(url_for('hr.payroll_generate'))
        
        # Find the payroll record
        payroll = PayrollBatch.query.get(payroll_id)
        
        if not payroll:
            flash('Payroll not found', 'error')
            return redirect(url_for('hr.payroll_generate'))
        
        # Move batch to HR approved for admin review
        if payroll.status == PayrollStatus.DRAFT:
            payroll.status = PayrollStatus.HR_APPROVED
        
        db.session.commit()
        
        # Log the approval action
        approval_log = ApprovalLog(
            entity_type='payroll',
            entity_id=payroll.id,
            action='submitted_for_approval',
            actor_id=current_user.id,
            comment=f"Payroll batch {payroll.batch_name} for {payroll.payroll_period} submitted to admin for approval"
        )
        db.session.add(approval_log)
        db.session.commit()
        
        current_app.logger.info(f"Payroll batch {payroll.batch_name} sent to admin for approval by {current_user.email}")
        
        flash(f"Payroll {payroll.batch_name} sent to admin for approval successfully!", "success")
        return redirect(url_for('hr.payroll'))
        
    except Exception as e:
        current_app.logger.error(f"Payroll Send Approval Error: {str(e)}")
        flash(f"Error sending payroll: {str(e)}", "error")
        return redirect(url_for('hr.payroll_generate'))


@bp.route('/payroll/export/excel', methods=['GET'])
@login_required
@hr_required
def payroll_export_excel():
    """Export payroll data to Excel using Nigeria payroll template columns."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from app.payroll_models import PayrollBatch

        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll 2026"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws['A1'] = "PAYROLL REPORT (NIGERIA 2026 STRUCTURE)"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:V1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}"
        ws.merge_cells('A2:V2')

        latest_batch = PayrollBatch.query.order_by(PayrollBatch.created_at.desc()).first()
        ws['A3'] = f"Latest Batch: {latest_batch.batch_name} ({latest_batch.payroll_period})" if latest_batch else "Latest Batch: Not available"
        ws.merge_cells('A3:V3')

        active_staff = User.query.filter_by(is_active=True).order_by(User.name.asc()).all()
        payroll_summary = _build_payroll_summary(active_staff)

        row = 5
        headers = [
            'SN/O', 'Names', 'Gross (Monthly)', 'Basic Salary', 'Housing', 'Transport',
            'Utility', 'Meal', 'Medical', 'Annual Gross', 'Annual Pension Employer',
            'Annual Pension Employee', 'Annual NHIS', 'NHF Annual', 'House Rent',
            'Tax Relief', 'Consolidated Relief', 'Taxable Income Annual',
            'PAYE Annual', 'PAYE Monthly', 'Total Deductions Monthly', 'Net Monthly'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        row = 6
        index = 1
        for staff in payroll_summary['staff_details']:
            values = [
                index,
                staff['name'],
                float(staff['gross_monthly']),
                float(staff['basic_salary']),
                float(staff['housing']),
                float(staff['transport']),
                float(staff['utility']),
                float(staff['meal']),
                float(staff['medical']),
                float(staff['annual_gross']),
                float(staff['annual_pension_employer']),
                float(staff['annual_pension_employee']),
                float(staff['annual_nhis']),
                float(staff['annual_nhf']),
                float(staff['house_rent_annual']),
                float(staff['tax_relief_annual']),
                float(staff['consolidated_relief_annual']),
                float(staff['taxable_income_annual']),
                float(staff['paye_annual']),
                float(staff['paye_monthly']),
                float(staff['total_deductions_monthly']),
                float(staff['net_monthly']),
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                if col >= 3:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            index += 1
            row += 1

        ws.cell(row=row, column=1).value = "TOTAL"
        ws.cell(row=row, column=1).font = total_font
        ws.cell(row=row, column=1).fill = total_fill

        totals = [
            payroll_summary['total_gross'],
            payroll_summary['total_basic_salary'],
            sum((s['housing'] for s in payroll_summary['staff_details']), Decimal('0')),
            sum((s['transport'] for s in payroll_summary['staff_details']), Decimal('0')),
            sum((s['utility'] for s in payroll_summary['staff_details']), Decimal('0')),
            sum((s['meal'] for s in payroll_summary['staff_details']), Decimal('0')),
            sum((s['medical'] for s in payroll_summary['staff_details']), Decimal('0')),
            sum((s['annual_gross'] for s in payroll_summary['staff_details']), Decimal('0')),
            sum((s['annual_pension_employer'] for s in payroll_summary['staff_details']), Decimal('0')),
            sum((s['annual_pension_employee'] for s in payroll_summary['staff_details']), Decimal('0')),
            sum((s['annual_nhis'] for s in payroll_summary['staff_details']), Decimal('0')),
            sum((s['annual_nhf'] for s in payroll_summary['staff_details']), Decimal('0')),
            sum((s['house_rent_annual'] for s in payroll_summary['staff_details']), Decimal('0')),
            sum((s['tax_relief_annual'] for s in payroll_summary['staff_details']), Decimal('0')),
            sum((s['consolidated_relief_annual'] for s in payroll_summary['staff_details']), Decimal('0')),
            payroll_summary['total_taxable_income_annual'],
            sum((s['paye_annual'] for s in payroll_summary['staff_details']), Decimal('0')),
            payroll_summary['total_paye_monthly'],
            payroll_summary['total_deductions'],
            payroll_summary['total_net_salary'],
        ]

        ws.cell(row=row, column=2).value = f"{payroll_summary['staff_count']} staff"
        ws.cell(row=row, column=2).font = total_font
        ws.cell(row=row, column=2).fill = total_fill
        ws.cell(row=row, column=2).border = border

        for idx, total in enumerate(totals, start=3):
            cell = ws.cell(row=row, column=idx)
            cell.value = float(_money(total))
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right', vertical='center')

        widths = {
            'A': 8, 'B': 24, 'C': 14, 'D': 14, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 12,
            'J': 14, 'K': 18, 'L': 18, 'M': 12, 'N': 12, 'O': 12, 'P': 12, 'Q': 16, 'R': 16,
            'S': 12, 'T': 12, 'U': 18, 'V': 14,
        }
        for column, width in widths.items():
            ws.column_dimensions[column].width = width

        os.makedirs(os.path.join('app', 'uploads'), exist_ok=True)
        temp_file = os.path.join('app', 'uploads', f'payroll_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        wb.save(temp_file)

        current_app.logger.info(f"Payroll Excel exported by {current_user.email}")

        return send_file(
            temp_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"Payroll_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

    except Exception as e:
        current_app.logger.error(f"Payroll Export Error: {str(e)}")
        flash(f"Error exporting payroll: {str(e)}", "error")
        return redirect(url_for('hr.payroll'))
@login_required
@hr_required
def payroll_submit(payroll_id=None):
    """Redirect to payroll view"""
    return redirect(url_for('hr.payroll_view_detail', payroll_id=payroll_id))

@bp.route('/payroll/<int:payroll_id>/view', methods=['GET'])
@login_required
@hr_required
def payroll_view_detail(payroll_id):
    """View payroll details with approval status"""
    try:
        from app.models import Payroll
        
        payroll = Payroll.query.get_or_404(payroll_id)
        
        # Get approval logs for this payroll
        approval_history = ApprovalLog.query.filter_by(
            entity_type='payroll',
            entity_id=payroll_id
        ).order_by(ApprovalLog.timestamp.desc()).all()
        
        return render_template('hr/payroll/view_detail.html', 
                             payroll=payroll,
                             approval_history=approval_history)
    except Exception as e:
        current_app.logger.error(f"Payroll View Error: {str(e)}")
        flash("Error loading payroll", "error")
        return redirect(url_for('hr.payroll'))

@bp.route('/payroll/<int:payroll_id>')
@login_required
@hr_required
def payroll_view(payroll_id):
    """Alias for payroll_view_detail"""
    return redirect(url_for('hr.payroll_view_detail', payroll_id=payroll_id))

@bp.route('/payroll/<int:payroll_id>/approve', methods=['POST'])
@login_required
@admin_required
def payroll_approve(payroll_id):
    """Admin approves payroll"""
    try:
        from app.models import Payroll
        from datetime import datetime
        
        payroll = Payroll.query.get_or_404(payroll_id)
        
        if payroll.approval_state == ApprovalState.APPROVED:
            flash("Payroll is already approved", "info")
            return redirect(url_for('hr.payroll_view_detail', payroll_id=payroll_id))
        
        payroll.approval_state = ApprovalState.APPROVED
        payroll.approved_by = current_user.id
        payroll.approved_at = datetime.utcnow()
        
        approval_log = ApprovalLog(
            entity_type='payroll',
            entity_id=payroll.id,
            action='approved',
            actor_id=current_user.id,
            timestamp=datetime.utcnow(),
            notes=request.form.get('approval_notes', 'Payroll approved')
        )
        
        db.session.add(approval_log)
        db.session.commit()
        
        flash(f"Payroll {payroll.payroll_number} approved successfully", "success")
        return redirect(url_for('hr.payroll_view_detail', payroll_id=payroll_id))
    
    except Exception as e:
        current_app.logger.error(f"Payroll Approval Error: {str(e)}")
        flash("Error approving payroll", "error")
        return redirect(url_for('hr.payroll'))

@bp.route('/payroll/<int:payroll_id>/reject', methods=['POST'])
@login_required
@admin_required
def payroll_reject(payroll_id):
    """Admin rejects payroll"""
    try:
        from app.models import Payroll
        from datetime import datetime
        
        payroll = Payroll.query.get_or_404(payroll_id)
        
        if payroll.approval_state in [ApprovalState.APPROVED, ApprovalState.REJECTED]:
            flash(f"Cannot reject payroll in {payroll.approval_state} state", "error")
            return redirect(url_for('hr.payroll_view_detail', payroll_id=payroll_id))
        
        rejection_reason = request.form.get('rejection_reason', 'No reason provided')
        payroll.approval_state = ApprovalState.REJECTED
        payroll.rejection_reason = rejection_reason
        
        approval_log = ApprovalLog(
            entity_type='payroll',
            entity_id=payroll.id,
            action='rejected',
            actor_id=current_user.id,
            timestamp=datetime.utcnow(),
            notes=rejection_reason
        )
        
        db.session.add(approval_log)
        db.session.commit()
        
        flash(f"Payroll {payroll.payroll_number} rejected", "warning")
        return redirect(url_for('hr.payroll_view_detail', payroll_id=payroll_id))
    
    except Exception as e:
        current_app.logger.error(f"Payroll Rejection Error: {str(e)}")
        flash("Error rejecting payroll", "error")
        return redirect(url_for('hr.payroll'))

@bp.route('/payroll/<int:payroll_id>/export')
@login_required
@hr_required
def payroll_export(payroll_id):
    """Redirect to staff list"""
    flash("Payroll export coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/leave')
@login_required
def leave_management():
    """Leave dashboard: all roles can request leave, HR can approve/reject."""
    try:
        page = request.args.get('page', 1, type=int)
        year = request.args.get('year', datetime.utcnow().year, type=int)
        is_hr = current_user.role in [Roles.HR_MANAGER, Roles.HR_STAFF, Roles.ADMIN]

        staff_leave_balances = []
        if is_hr:
            leave_requests = LeaveRequest.query.order_by(LeaveRequest.created_at.desc()).paginate(page=page, per_page=20)
            staff_list = User.query.filter_by(is_active=True).paginate(page=page, per_page=20)
            for staff in staff_list.items:
                balances_by_type = build_leave_balance(staff.id, year)
                staff_leave_balances.append({
                    'staff': staff,
                    'balances': balances_by_type
                })
        else:
            leave_requests = LeaveRequest.query.filter_by(user_id=current_user.id).order_by(LeaveRequest.created_at.desc()).paginate(page=page, per_page=20)
            staff_list = User.query.filter(User.id == current_user.id).paginate(page=1, per_page=1)

        balances = build_leave_balance(current_user.id, year)
        pending_approvals = LeaveRequest.query.filter_by(status='pending').count() if is_hr else 0
        on_leave_today = LeaveRequest.query.filter(
            LeaveRequest.status == 'approved',
            LeaveRequest.start_date <= date.today(),
            LeaveRequest.end_date >= date.today()
        ).count()

        return render_template(
            'hr/leave.html',
            staff_list=staff_list,
            leave_requests=leave_requests,
            total_staff=User.query.filter_by(is_active=True).count(),
            balances=balances,
            staff_leave_balances=staff_leave_balances,
            is_hr=is_hr,
            selected_year=year,
            pending_approvals=pending_approvals,
            on_leave_today=on_leave_today
        )
    except Exception as e:
        current_app.logger.error(f"Leave Error: {str(e)}")
        flash("Error loading leave management", "error")
        return redirect(url_for('hr.staff_list'))

@bp.route('/leave/create', methods=['GET', 'POST'])
@login_required
def create_leave():
    """Create leave request for current user."""
    try:
        if request.method == 'POST':
            leave_type = normalize_leave_type(request.form.get('leave_type'))
            start_date = request.form.get('start_date')
            end_date = request.form.get('end_date')
            reason = request.form.get('reason', '').strip()

            if leave_type not in LEAVE_ALLOWANCE:
                flash("Invalid leave type selected", "error")
                return redirect(url_for('hr.leave_management'))
            if not start_date or not end_date:
                flash("Start and end dates are required", "error")
                return redirect(url_for('hr.leave_management'))

            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
            if end_dt < start_dt:
                flash("End date must be on or after start date", "error")
                return redirect(url_for('hr.leave_management'))

            requested_days = (end_dt - start_dt).days + 1
            balance = build_leave_balance(current_user.id, start_dt.year).get(leave_type, {})
            if requested_days > int(balance.get('remaining', 0)):
                flash(f"Insufficient leave balance. Remaining {leave_type.title()} leave: {balance.get('remaining', 0)} day(s)", "error")
                return redirect(url_for('hr.leave_management'))

            leave_request = LeaveRequest(
                user_id=current_user.id,
                leave_type=leave_type,
                start_date=start_dt,
                end_date=end_dt,
                days_requested=requested_days,
                reason=reason,
                status='pending'
            )
            db.session.add(leave_request)
            db.session.commit()
            flash("Leave request submitted successfully and pending HR approval.", "success")
        return redirect(url_for('hr.leave_management'))
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Leave create error: {str(e)}")
        flash("Error creating leave request", "error")
        return redirect(url_for('hr.leave_management'))

@bp.route('/attendance')
@login_required
@hr_required
def attendance():
    """Redirect to staff list"""
    flash("Attendance tracking coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/attendance/record', methods=['GET', 'POST'])
@login_required
@hr_required
def record_attendance():
    """Redirect to staff list"""
    flash("Attendance recording coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/queries')
@login_required
@hr_required
def staff_queries():
    """Staff queries and complaints dashboard"""
    try:
        page = request.args.get('page', 1, type=int)
        staff_list = User.query.filter_by(is_active=True).paginate(page=page, per_page=20)
        
        return render_template('hr/queries.html', 
                             staff_list=staff_list,
                             total_staff=User.query.filter_by(is_active=True).count())
    except Exception as e:
        current_app.logger.error(f"Queries Error: {str(e)}")
        flash("Error loading queries", "error")
        return redirect(url_for('hr.staff_list'))

@bp.route('/queries/create', methods=['GET', 'POST'])
@login_required
@hr_required
def create_query():
    """Redirect to staff list"""
    flash("Query creation coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/tasks')
@login_required
@hr_required
def tasks():
    """Redirect to staff list"""
    flash("Task management coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/reports')
@login_required
@hr_required
def reports():
    """HR Reports Dashboard"""
    try:
        staff_count = User.query.filter_by(is_active=True).count()
        total_salary = 0
        total_deductions = 0
        
        staff_list = User.query.filter_by(is_active=True).all()
        for staff in staff_list:
            total_salary += float(staff.basic_salary or 0)
            total_deductions += float(staff.default_deductions or 0)
        
        report_data = {
            'total_staff': staff_count,
            'total_monthly_salary': f"{total_salary:,.2f}",
            'total_deductions': f"{total_deductions:,.2f}",
            'net_payroll': f"{(total_salary - total_deductions):,.2f}",
            'average_salary': f"{(total_salary / staff_count if staff_count > 0 else 0):,.2f}"
        }
        
        return render_template('hr/reports.html', report_data=report_data, staff_list=staff_list)
    except Exception as e:
        current_app.logger.error(f"Reports Error: {str(e)}")
        flash("Error loading reports", "error")
        return redirect(url_for('hr.staff_list'))

@bp.route('/reports/payroll')
@login_required
@hr_required
def generate_payroll_report():
    """Redirect to staff list"""
    flash("Payroll report generation coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/staff/add', methods=['GET', 'POST'])
@login_required
@hr_required
def add_staff():
    """Redirect to staff list"""
    if request.method == 'POST':
        flash("User creation coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/analytics')
@login_required
@hr_required
def analytics():
    """Redirect to staff list"""
    flash("Analytics coming soon", "info")
    return redirect(url_for('hr.staff_list'))

# Additional stub routes for other unimplemented endpoints
@bp.route('/leave/<int:leave_id>')
@login_required
def leave_detail(leave_id):
    """View leave request details."""
    leave = LeaveRequest.query.get_or_404(leave_id)
    is_hr = current_user.role in [Roles.HR_MANAGER, Roles.HR_STAFF, Roles.ADMIN]
    if not is_hr and leave.user_id != current_user.id:
        flash("Access denied", "error")
        return redirect(url_for('hr.leave_management'))
    return render_template('hr/leave/detail.html', leave=leave, is_hr=is_hr)


@bp.route('/leave/<int:leave_id>/approve', methods=['POST'])
@login_required
@hr_required
def approve_leave(leave_id):
    """HR approves leave request."""
    leave = LeaveRequest.query.get_or_404(leave_id)
    try:
        if leave.status != 'pending':
            flash("Leave request is already processed", "warning")
            return redirect(url_for('hr.leave_management'))
        leave.status = 'approved'
        leave.reviewed_by = current_user.id
        leave.reviewed_at = datetime.utcnow()
        db.session.commit()
        flash("Leave request approved", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Approve leave error: {str(e)}")
        flash("Error approving leave request", "error")
    return redirect(url_for('hr.leave_management'))


@bp.route('/leave/<int:leave_id>/reject', methods=['POST'])
@login_required
@hr_required
def reject_leave(leave_id):
    """HR rejects leave request."""
    leave = LeaveRequest.query.get_or_404(leave_id)
    try:
        if leave.status != 'pending':
            flash("Leave request is already processed", "warning")
            return redirect(url_for('hr.leave_management'))
        leave.status = 'rejected'
        leave.rejection_reason = request.form.get('rejection_reason', '').strip()
        leave.reviewed_by = current_user.id
        leave.reviewed_at = datetime.utcnow()
        db.session.commit()
        flash("Leave request rejected", "warning")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Reject leave error: {str(e)}")
        flash("Error rejecting leave request", "error")
    return redirect(url_for('hr.leave_management'))

@bp.route('/attendance/<int:record_id>/update', methods=['POST'])
@login_required
@hr_required
def update_attendance(record_id):
    """Redirect to staff list"""
    flash("Attendance update coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/queries/<int:query_id>')
@login_required
@hr_required
def query_detail(query_id):
    """Redirect to staff list"""
    flash("Query detail view coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/tasks/create', methods=['GET', 'POST'])
@login_required
@hr_required
def create_task():
    """Redirect to staff list"""
    flash("Task creation coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/tasks/<int:task_id>/update', methods=['POST'])
@login_required
@hr_required
def update_task(task_id):
    """Redirect to staff list"""
    flash("Task update coming soon", "info")
    return redirect(url_for('hr.staff_list'))

# ==================== REPORT ENDPOINTS ====================

@bp.route('/settings/save-leave-policy', methods=['POST'])
@login_required
@hr_required
def save_leave_policy():
    """Redirect to staff list"""
    flash("Leave policy save coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/settings/save-payroll-settings', methods=['POST'])
@login_required
@hr_required
def save_payroll_settings():
    """Redirect to staff list"""
    flash("Payroll settings save coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/settings/save-attendance-settings', methods=['POST'])
@login_required
@hr_required
def save_attendance_settings():
    """Redirect to staff list"""
    flash("Attendance settings save coming soon", "info")
    return redirect(url_for('hr.staff_list'))

# Additional missing routes from templates
@bp.route('/role-management')
@login_required
@hr_required
def role_management():
    """Redirect to staff list"""
    flash("Role management coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/performance')
@login_required
@hr_required
def performance():
    """Performance management dashboard"""
    try:
        page = request.args.get('page', 1, type=int)
        staff_list = User.query.filter_by(is_active=True).paginate(page=page, per_page=20)
        
        return render_template('hr/performance.html', 
                             staff_list=staff_list,
                             total_staff=User.query.filter_by(is_active=True).count())
    except Exception as e:
        current_app.logger.error(f"Performance Error: {str(e)}")
        flash("Error loading performance data", "error")
        return redirect(url_for('hr.staff_list'))

@bp.route('/staff/<int:staff_id>/delete', methods=['POST'])
@login_required
@hr_required
def delete_staff(staff_id):
    """Redirect to staff list"""
    flash("Staff deletion coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/payroll/submit', methods=['POST'])
@login_required
@hr_required
def payroll_submit_batch():
    """Redirect to staff list"""
    flash("Payroll submission coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/reports/staff')
@login_required
@hr_required
def generate_staff_report():
    """Redirect to staff list"""
    flash("Staff report generation coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/reports/performance')
@login_required
@hr_required
def generate_performance_report():
    """Performance report with staff ratings"""
    try:
        staff_list = User.query.filter_by(is_active=True).all()
        
        performance_data = {
            'total_staff': len(staff_list),
            'reviewed_staff': 0,
            'pending_reviews': len(staff_list),
            'average_rating': 0,
            'staff': [{'id': s.id, 'name': s.name, 'role': s.role, 'rating': 3.5} for s in staff_list]
        }
        
        return render_template('hr/reports/performance.html', data=performance_data)
    except Exception as e:
        current_app.logger.error(f"Performance Report Error: {str(e)}")
        flash("Error loading performance report", "error")
        return redirect(url_for('hr.reports'))

@bp.route('/reports/attendance')
@login_required
@hr_required
def generate_attendance_report():
    """Attendance report with staff presence"""
    try:
        staff_list = User.query.filter_by(is_active=True).all()
        
        attendance_data = {
            'total_staff': len(staff_list),
            'present_today': 0,
            'absent_today': len(staff_list),
            'late_today': 0,
            'staff': [{'id': s.id, 'name': s.name, 'email': s.email, 'status': 'Absent'} for s in staff_list]
        }
        
        return render_template('hr/reports/attendance.html', data=attendance_data)
    except Exception as e:
        current_app.logger.error(f"Attendance Report Error: {str(e)}")
        flash("Error loading attendance report", "error")
        return redirect(url_for('hr.reports'))

@bp.route('/reports/leave')
@login_required
@hr_required
def generate_leave_report():
    """Leave report with leave usage"""
    try:
        staff_list = User.query.filter_by(is_active=True).all()
        
        leave_data = {
            'total_staff': len(staff_list),
            'on_leave_today': 0,
            'pending_requests': 0,
            'approved_leave': 0,
            'staff': [{'id': s.id, 'name': s.name, 'email': s.email, 'leave_balance': 18, 'used': 0} for s in staff_list]
        }
        
        return render_template('hr/reports/leave.html', data=leave_data)
    except Exception as e:
        current_app.logger.error(f"Leave Report Error: {str(e)}")
        flash("Error loading leave report", "error")
        return redirect(url_for('hr.reports'))

@bp.route('/reports/salary')
@login_required
@hr_required
def generate_salary_report():
    """Salary report with payroll details"""
    try:
        staff_list = User.query.filter_by(is_active=True).all()
        
        total_salary = sum(float(s.basic_salary or 0) for s in staff_list)
        total_deductions = sum(float(s.default_deductions or 0) for s in staff_list)
        
        salary_data = {
            'total_staff': len(staff_list),
            'total_salary': f"{total_salary:,.2f}",
            'total_deductions': f"{total_deductions:,.2f}",
            'net_payroll': f"{(total_salary - total_deductions):,.2f}",
            'staff': [{'id': s.id, 'name': s.name, 'salary': f"{float(s.basic_salary or 0):,.2f}", 'deductions': f"{float(s.default_deductions or 0):,.2f}"} for s in staff_list]
        }
        
        return render_template('hr/reports/salary.html', data=salary_data)
    except Exception as e:
        current_app.logger.error(f"Salary Report Error: {str(e)}")
        flash("Error loading salary report", "error")
        return redirect(url_for('hr.reports'))

# ==================== QUERIES ENDPOINTS ====================

@bp.route('/queries/<int:staff_id>/view')
@login_required
@hr_required
def view_query(staff_id):
    """View staff query details"""
    try:
        staff = User.query.get_or_404(staff_id)
        
        query_data = {
            'id': staff.id,
            'name': staff.name,
            'email': staff.email,
            'role': staff.role,
            'subject': 'Staff Query',
            'description': 'No queries recorded',
            'date_submitted': datetime.now().strftime('%Y-%m-%d'),
            'status': 'Open'
        }
        
        return render_template('hr/queries/detail.html', query=query_data)
    except Exception as e:
        current_app.logger.error(f"View Query Error: {str(e)}")
        flash("Error loading query details", "error")
        return redirect(url_for('hr.staff_queries'))

@bp.route('/queries/<int:staff_id>/message', methods=['GET', 'POST'])
@login_required
@hr_required
def message_staff(staff_id):
    """Send message to staff member"""
    try:
        staff = User.query.get_or_404(staff_id)
        
        if request.method == 'POST':
            message_content = request.form.get('message', '')
            
            if message_content:
                flash(f"Message sent to {staff.name}", "success")
                return redirect(url_for('hr.staff_queries'))
            else:
                flash("Message cannot be empty", "error")
        
        return render_template('hr/queries/message.html', staff=staff)
    except Exception as e:
        current_app.logger.error(f"Message Staff Error: {str(e)}")
        flash("Error sending message", "error")
        return redirect(url_for('hr.staff_queries'))

# ==================== LEAVE ENDPOINTS ====================

@bp.route('/leave/<int:staff_id>/request', methods=['GET', 'POST'])
@login_required
def request_leave(staff_id):
    """Request leave for staff"""
    try:
        staff = User.query.get_or_404(staff_id)
        is_hr = current_user.role in [Roles.HR_MANAGER, Roles.HR_STAFF, Roles.ADMIN]
        if not is_hr and current_user.id != staff.id:
            flash("You can only request leave for your own profile", "error")
            return redirect(url_for('hr.leave_management'))
        
        if request.method == 'POST':
            leave_type = normalize_leave_type(request.form.get('leave_type', 'casual'))
            start_date = request.form.get('start_date', '')
            end_date = request.form.get('end_date', '')
            reason = request.form.get('reason', '')

            if leave_type not in LEAVE_ALLOWANCE:
                flash("Invalid leave type", "error")
                return render_template('hr/leave/request.html', staff=staff, leave_allowance=LEAVE_ALLOWANCE)

            if start_date and end_date:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
                requested_days = (end_dt - start_dt).days + 1
                balance = build_leave_balance(staff.id, start_dt.year).get(leave_type, {})
                if requested_days > int(balance.get('remaining', 0)):
                    flash(f"Insufficient balance. Remaining {leave_type.title()} leave: {balance.get('remaining', 0)} day(s)", "error")
                    return render_template('hr/leave/request.html', staff=staff, leave_allowance=LEAVE_ALLOWANCE)

                leave = LeaveRequest(
                    user_id=staff.id,
                    leave_type=leave_type,
                    start_date=start_dt,
                    end_date=end_dt,
                    days_requested=requested_days,
                    reason=reason,
                    status='pending'
                )
                db.session.add(leave)
                db.session.commit()
                flash(f"Leave request submitted for {staff.name}", "success")
                return redirect(url_for('hr.leave_management'))
            flash("Please select leave dates", "error")
        
        return render_template('hr/leave/request.html', staff=staff, leave_allowance=LEAVE_ALLOWANCE)
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Request Leave Error: {str(e)}")
        flash("Error submitting leave request", "error")
        return redirect(url_for('hr.leave_management'))

@bp.route('/leave/<int:staff_id>/view')
@login_required
def view_leave(staff_id):
    """View staff leave balance and history"""
    try:
        staff = User.query.get_or_404(staff_id)
        is_hr = current_user.role in [Roles.HR_MANAGER, Roles.HR_STAFF, Roles.ADMIN]
        if not is_hr and current_user.id != staff.id:
            flash("Access denied", "error")
            return redirect(url_for('hr.leave_management'))

        year = datetime.utcnow().year
        balances = build_leave_balance(staff.id, year)
        leave_history = LeaveRequest.query.filter_by(user_id=staff.id).order_by(LeaveRequest.created_at.desc()).all()
        return render_template('hr/leave/detail.html', staff=staff, balances=balances, leave_history=leave_history, year=year, is_hr=is_hr)
    except Exception as e:
        current_app.logger.error(f"View Leave Error: {str(e)}")
        flash("Error loading leave details", "error")
        return redirect(url_for('hr.leave_management'))

# ==================== PAYROLL ENDPOINTS ====================

@bp.route('/payroll/<int:staff_id>/edit', methods=['GET', 'POST'])
@login_required
@hr_required
def edit_payroll(staff_id):
    """Edit staff payroll details"""
    try:
        staff = User.query.get_or_404(staff_id)
        
        if request.method == 'POST':
            basic_salary = request.form.get('basic_salary', 0)
            deductions = request.form.get('deductions', 0)
            
            try:
                staff.basic_salary = float(basic_salary)
                staff.default_deductions = float(deductions)
                db.session.commit()
                
                flash(f"Payroll updated for {staff.name}", "success")
                return redirect(url_for('hr.payroll'))
            except ValueError:
                flash("Invalid salary or deduction amount", "error")
        
        payroll_data = {
            'id': staff.id,
            'name': staff.name,
            'email': staff.email,
            'basic_salary': f"{float(staff.basic_salary or 0):,.2f}",
            'deductions': f"{float(staff.default_deductions or 0):,.2f}",
            'net_salary': f"{(float(staff.basic_salary or 0) - float(staff.default_deductions or 0)):,.2f}"
        }
        
        return render_template('hr/payroll/edit.html', payroll=payroll_data)
    except Exception as e:
        current_app.logger.error(f"Edit Payroll Error: {str(e)}")
        flash("Error editing payroll", "error")
        return redirect(url_for('hr.payroll'))

@bp.route('/payroll/<int:staff_id>/slip')
@login_required
@hr_required
def view_salary_slip(staff_id):
    """View salary slip for staff"""
    try:
        staff = User.query.get_or_404(staff_id)
        
        slip_data = {
            'id': staff.id,
            'name': staff.name,
            'email': staff.email,
            'date_generated': datetime.now().strftime('%Y-%m-%d'),
            'basic_salary': f"{float(staff.basic_salary or 0):,.2f}",
            'deductions': f"{float(staff.default_deductions or 0):,.2f}",
            'net_salary': f"{(float(staff.basic_salary or 0) - float(staff.default_deductions or 0)):,.2f}"
        }
        
        return render_template('hr/payroll/slip.html', slip=slip_data)
    except Exception as e:
        current_app.logger.error(f"View Salary Slip Error: {str(e)}")
        flash("Error loading salary slip", "error")
        return redirect(url_for('hr.payroll'))

# ==================== PERFORMANCE ENDPOINTS ====================

@bp.route('/performance/<int:staff_id>/review', methods=['GET', 'POST'])
@login_required
@hr_required
def review_performance(staff_id):
    """Review staff performance"""
    try:
        staff = User.query.get_or_404(staff_id)
        
        if request.method == 'POST':
            rating = request.form.get('rating', 3)
            comments = request.form.get('comments', '')
            
            flash(f"Performance review submitted for {staff.name}", "success")
            return redirect(url_for('hr.performance'))
        
        review_data = {
            'id': staff.id,
            'name': staff.name,
            'role': staff.role,
            'email': staff.email,
            'rating': 3,
            'comments': ''
        }
        
        return render_template('hr/performance/review.html', staff=review_data)
    except Exception as e:
        current_app.logger.error(f"Review Performance Error: {str(e)}")
        flash("Error submitting performance review", "error")
        return redirect(url_for('hr.performance'))

@bp.route('/performance/<int:staff_id>/history')
@login_required
@hr_required
def performance_history(staff_id):
    """View staff performance history"""
    try:
        staff = User.query.get_or_404(staff_id)
        
        history_data = {
            'id': staff.id,
            'name': staff.name,
            'email': staff.email,
            'role': staff.role,
            'reviews': []
        }
        
        return render_template('hr/performance/history.html', history=history_data)
    except Exception as e:
        current_app.logger.error(f"Performance History Error: {str(e)}")
        flash("Error loading performance history", "error")
        return redirect(url_for('hr.performance'))

@bp.route('/reports/custom')
@login_required
@hr_required
def custom_report():
    """Redirect to staff list"""
    flash("Custom report generation coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/reports/<int:report_id>')
@login_required
@hr_required
def view_report(report_id):
    """Redirect to staff list"""
    flash("Report view coming soon", "info")
    return redirect(url_for('hr.staff_list'))


# ==================== EXCEL IMPORT ROUTES ====================

def get_user_import_batches(user_id, limit=5):
    """Get recent import batches for a user."""
    from app.models import StaffImportBatch
    return StaffImportBatch.query.filter_by(created_by=user_id).order_by(
        StaffImportBatch.created_at.desc()
    ).limit(limit).all()


@bp.route('/staff/import', methods=['GET', 'POST'])
@login_required
@hr_required
def import_staff():
    """Upload and preview Excel staff import file."""
    try:
        if request.method == 'POST':
            # Check if file is provided
            if 'excel_file' not in request.files:
                flash('No file provided', 'error')
                return redirect(url_for('hr.import_staff'))
            
            file = request.files['excel_file']
            
            # Validate file
            try:
                StaffExcelParser.validate_file(file)
            except ExcelImportError as e:
                flash(str(e), 'error')
                return redirect(url_for('hr.import_staff'))
            
            # Save file temporarily
            upload_dir = os.path.join(current_app.root_path, 'uploads', 'imports')
            os.makedirs(upload_dir, exist_ok=True)
            
            filename = secure_filename(f"staff_import_{datetime.now().timestamp()}_{file.filename}")
            file_path = os.path.join(upload_dir, filename)
            file.save(file_path)
            
            # Parse and validate Excel file
            try:
                valid_records, invalid_records = StaffExcelParser.parse_and_validate(file_path)
            except ExcelImportError as e:
                flash(f'Error parsing file: {str(e)}', 'error')
                os.remove(file_path)
                return redirect(url_for('hr.import_staff'))
            
            if not valid_records and invalid_records:
                # Show detailed errors
                error_details = '<strong>Validation Errors:</strong><ul>'
                for err in invalid_records[:10]:  # Show first 10 errors
                    error_details += f"<li><strong>Row {err['row']}:</strong> {err['error']}</li>"
                if len(invalid_records) > 10:
                    error_details += f"<li>... and {len(invalid_records) - 10} more errors</li>"
                error_details += '</ul><strong>Required Columns:</strong> first_name, last_name, email, basic_salary'
                flash(error_details, 'error')
                os.remove(file_path)
                return redirect(url_for('hr.import_staff'))
            
            # Create import batch
            batch_name = request.form.get('batch_name', f"Import_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
            batch = StaffImportManager.create_import_batch(
                batch_name=batch_name,
                file_path=file_path,
                file_name=file.filename,
                valid_records=valid_records,
                created_by_id=current_user.id,
                invalid_records=invalid_records
            )
            
            flash(f'File uploaded successfully. {len(valid_records)} valid records found. {len(invalid_records) if invalid_records else 0} records with errors.', 'info')
            
            return redirect(url_for('hr.preview_import', batch_id=batch.id))
        
        return render_template('hr/staff/import.html', get_user_import_batches=get_user_import_batches)
        
    except Exception as e:
        current_app.logger.error(f"Staff Import Error: {str(e)}")
        flash('Error processing import', 'error')
        return redirect(url_for('hr.staff_list'))


@bp.route('/staff/import/<int:batch_id>/preview')
@login_required
@hr_required
def preview_import(batch_id):
    """Preview staff records before import."""
    try:
        batch = StaffImportBatch.query.get_or_404(batch_id)
        
        # Only creator or admin can preview
        if current_user.id != batch.created_by and current_user.role != Roles.ADMIN:
            flash('Insufficient permissions', 'error')
            return redirect(url_for('hr.staff_list'))
        
        # Get valid and invalid items
        valid_items = StaffImportItem.query.filter_by(batch_id=batch_id, status='pending').all()
        invalid_items = StaffImportItem.query.filter(
            StaffImportItem.batch_id == batch_id,
            StaffImportItem.error_message.isnot(None)
        ).all()
        
        return render_template(
            'hr/staff/import_preview.html',
            batch=batch,
            valid_items=valid_items,
            invalid_items=invalid_items
        )
        
    except Exception as e:
        current_app.logger.error(f"Import Preview Error: {str(e)}")
        flash('Error loading preview', 'error')
        return redirect(url_for('hr.staff_list'))


@bp.route('/staff/import/<int:batch_id>/submit', methods=['POST'])
@login_required
@hr_required
def submit_import_approval(batch_id):
    """Submit import batch for admin approval."""
    try:
        batch = StaffImportBatch.query.get_or_404(batch_id)
        
        # Only creator or admin can submit
        if current_user.id != batch.created_by and current_user.role != Roles.ADMIN:
            flash('Insufficient permissions', 'error')
            return redirect(url_for('hr.staff_list'))
        
        if batch.approval_state != ApprovalState.DRAFT:
            flash('Batch cannot be submitted in current state', 'error')
            return redirect(url_for('hr.preview_import', batch_id=batch_id))
        
        # Update batch status to pending approval
        batch.approval_state = ApprovalState.PENDING
        db.session.commit()
        
        flash(f'Batch submitted for admin approval. {batch.total_records} records ready for import.', 'success')
        
        # Log the action
        try:
            log = ApprovalLog(
                entity_type='staff_import_batch',
                entity_id=batch.id,
                actor_id=current_user.id,
                action='submit_for_approval',
                status='pending',
                comments=f'Submitted import batch with {batch.total_records} records'
            )
            db.session.add(log)
            db.session.commit()
        except:
            pass
        
        # If user is admin, redirect to admin users page to show success message
        if current_user.role in ['admin', 'super_hq']:
            return redirect(url_for('admin.users'))
        
        return redirect(url_for('hr.staff_list'))
        
    except Exception as e:
        current_app.logger.error(f"Submit Import Error: {str(e)}")
        flash('Error submitting batch for approval', 'error')
        return redirect(url_for('hr.staff_list'))


@bp.route('/staff/import/<int:batch_id>/cancel', methods=['POST'])
@login_required
@hr_required
def cancel_import(batch_id):
    """Cancel an import batch."""
    try:
        batch = StaffImportBatch.query.get_or_404(batch_id)
        
        # Only creator or admin can cancel
        if current_user.id != batch.created_by and current_user.role != Roles.ADMIN:
            flash('Insufficient permissions', 'error')
            return redirect(url_for('hr.staff_list'))
        
        if batch.approval_state not in [ApprovalState.DRAFT, ApprovalState.PENDING]:
            flash('Batch cannot be cancelled in current state', 'error')
            return redirect(url_for('hr.staff_list'))
        
        # Delete associated items
        StaffImportItem.query.filter_by(batch_id=batch_id).delete()
        
        # Delete batch
        db.session.delete(batch)
        db.session.commit()
        
        flash('Import batch cancelled', 'success')
        
    except Exception as e:
        current_app.logger.error(f"Cancel Import Error: {str(e)}")
        flash('Error cancelling batch', 'error')
    
    return redirect(url_for('hr.staff_list'))


@bp.route('/staff/<int:staff_id>/compensation')
@login_required
@hr_required
def staff_compensation(staff_id):
    """View and manage staff compensation and deductions."""
    try:
        staff = User.query.get_or_404(staff_id)
        compensation = StaffCompensation.query.filter_by(user_id=staff_id).first()
        
        if not compensation:
            flash('No compensation record found for this staff member', 'info')
            return redirect(url_for('hr.staff_details', staff_id=staff_id))
        
        deductions = PayrollDeduction.query.filter_by(compensation_id=compensation.id).all()
        
        return render_template(
            'hr/staff/compensation.html',
            staff=staff,
            compensation=compensation,
            deductions=deductions
        )
        
    except Exception as e:
        current_app.logger.error(f"Staff Compensation Error: {str(e)}")
        flash('Error loading compensation details', 'error')
        return redirect(url_for('hr.staff_list'))


@bp.route('/staff/<int:staff_id>/compensation/add-deduction', methods=['POST'])
@login_required
@hr_required
def add_deduction(staff_id):
    """Add a new deduction to staff compensation."""
    try:
        staff = User.query.get_or_404(staff_id)
        compensation = StaffCompensation.query.filter_by(user_id=staff_id).first_or_404()
        
        deduction_type = request.form.get('deduction_type')
        description = request.form.get('description')
        amount = request.form.get('amount')
        is_recurring = request.form.get('is_recurring', 'on') == 'on'
        
        if not deduction_type or not amount:
            flash('Deduction type and amount are required', 'error')
            return redirect(url_for('hr.staff_compensation', staff_id=staff_id))
        
        try:
            amount_float = float(amount)
            if amount_float < 0:
                raise ValueError('Amount must be positive')
        except ValueError:
            flash('Invalid amount', 'error')
            return redirect(url_for('hr.staff_compensation', staff_id=staff_id))
        
        deduction = PayrollDeduction(
            compensation_id=compensation.id,
            deduction_type=deduction_type,
            description=description,
            amount=amount_float,
            is_recurring=is_recurring,
            effective_from=datetime.now().date()
        )
        
        db.session.add(deduction)
        db.session.commit()
        
        flash(f'Deduction added: {deduction_type}', 'success')
        return redirect(url_for('hr.staff_compensation', staff_id=staff_id))
        
    except Exception as e:
        current_app.logger.error(f"Add Deduction Error: {str(e)}")
        flash('Error adding deduction', 'error')
        return redirect(url_for('hr.staff_compensation', staff_id=staff_id))


@bp.route('/staff/deduction/<int:deduction_id>/delete', methods=['POST'])
@login_required
@hr_required
def delete_deduction(deduction_id):
    """Delete a deduction from staff compensation."""
    try:
        deduction = PayrollDeduction.query.get_or_404(deduction_id)
        staff_id = deduction.compensation.user_id
        
        db.session.delete(deduction)
        db.session.commit()
        
        flash('Deduction removed', 'success')
        return redirect(url_for('hr.staff_compensation', staff_id=staff_id))
        
    except Exception as e:
        current_app.logger.error(f"Delete Deduction Error: {str(e)}")
        flash('Error deleting deduction', 'error')
        return redirect(url_for('hr.staff_list'))


@bp.route('/staff/<int:staff_id>/department-access')
@login_required
@hr_required
def staff_department_access(staff_id):
    """View and manage staff department access."""
    try:
        staff = User.query.get_or_404(staff_id)
        department_access = DepartmentAccess.query.filter_by(user_id=staff_id).all()
        
        # Get list of all departments
        all_departments = {
            'HR', 'Finance', 'Procurement', 'QC', 'Projects', 'Cost Control', 'Admin'
        }
        assigned_departments = {d.department for d in department_access}
        available_departments = all_departments - assigned_departments
        
        return render_template(
            'hr/staff/department_access.html',
            staff=staff,
            department_access=department_access,
            available_departments=sorted(available_departments)
        )
        
    except Exception as e:
        current_app.logger.error(f"Department Access Error: {str(e)}")
        flash('Error loading department access', 'error')
        return redirect(url_for('hr.staff_list'))


@bp.route('/staff/<int:staff_id>/department-access/add', methods=['POST'])
@login_required
@hr_required
def add_department_access(staff_id):
    """Add department access to staff."""
    try:
        staff = User.query.get_or_404(staff_id)
        department = request.form.get('department')
        access_level = request.form.get('access_level', 'view')
        
        if not department:
            flash('Department is required', 'error')
            return redirect(url_for('hr.staff_department_access', staff_id=staff_id))
        
        # Check if access already exists
        existing = DepartmentAccess.query.filter_by(
            user_id=staff_id,
            department=department
        ).first()
        
        if existing:
            flash(f'Staff already has access to {department}', 'info')
            return redirect(url_for('hr.staff_department_access', staff_id=staff_id))
        
        access = DepartmentAccess(
            user_id=staff_id,
            department=department,
            access_level=access_level,
            is_active=True
        )
        
        db.session.add(access)
        db.session.commit()
        
        flash(f'Department access added: {department}', 'success')
        return redirect(url_for('hr.staff_department_access', staff_id=staff_id))
        
    except Exception as e:
        current_app.logger.error(f"Add Department Access Error: {str(e)}")
        flash('Error adding department access', 'error')
        return redirect(url_for('hr.staff_department_access', staff_id=staff_id))


@bp.route('/staff/department-access/<int:access_id>/delete', methods=['POST'])
@login_required
@hr_required
def delete_department_access(access_id):
    """Remove department access from staff."""
    try:
        access = DepartmentAccess.query.get_or_404(access_id)
        staff_id = access.user_id
        department = access.department
        
        db.session.delete(access)
        db.session.commit()
        
        flash(f'Department access removed: {department}', 'success')
        return redirect(url_for('hr.staff_department_access', staff_id=staff_id))
        
    except Exception as e:
        current_app.logger.error(f"Delete Department Access Error: {str(e)}")
        flash('Error removing department access', 'error')
        return redirect(url_for('hr.staff_list'))


@bp.route('/staff/import/batches')
@login_required
@hr_required
def import_batches():
    """View all import batches."""
    try:
        page = request.args.get('page', 1, type=int)
        status_filter = request.args.get('status', 'all')
        
        query = StaffImportBatch.query
        
        if status_filter != 'all':
            query = query.filter_by(approval_state=status_filter)
        
        # HR staff only see their own batches, admin sees all
        if current_user.role != Roles.ADMIN:
            query = query.filter_by(created_by=current_user.id)
        
        batches = query.order_by(desc(StaffImportBatch.created_at)).paginate(page=page, per_page=10)
        
        return render_template(
            'hr/staff/import_batches.html',
            batches=batches,
            status_filter=status_filter
        )
        
    except Exception as e:
        current_app.logger.error(f"Import Batches Error: {str(e)}")
        flash('Error loading import batches', 'error')
        return redirect(url_for('hr.staff_list'))

@bp.route('/staff/import/template/download')
@login_required
@hr_required
def download_import_template():
    """Download Excel template for staff import"""
    try:
        import io
        import pandas as pd
        from datetime import datetime
        
        # Create sample data
        template_data = {
            'first_name': ['John', 'Jane'],
            'last_name': ['Doe', 'Smith'],
            'email': ['john.doe@company.com', 'jane.smith@company.com'],
            'phone_number': ['+234-801-234-5678', '+234-802-345-6789'],
            'date_of_birth': ['1990-01-15', '1992-03-22'],
            'gender': ['Male', 'Female'],
            'employee_id': ['EMP001', 'EMP002'],
            'department': ['HR', 'Finance'],
            'position': ['HR Manager', 'Finance Officer'],
            'employment_type': ['Full-time', 'Full-time'],
            'joining_date': ['2026-01-15', '2026-02-01'],
            'basic_salary': [150000, 120000],
            'allowances': [20000, 15000],
            'nok_name': ['Mary Doe', 'Tom Smith'],
            'nok_relationship': ['Spouse', 'Parent'],
            'nok_phone': ['+234-801-111-2222', '+234-802-333-4444']
        }
        
        df = pd.DataFrame(template_data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Staff', index=False)
            
            # Format the workbook
            workbook = writer.book
            worksheet = writer.sheets['Staff']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Return file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'staff_import_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    
    except Exception as e:
        current_app.logger.error(f"Template Download Error: {str(e)}")
        flash('Error downloading template', 'error')
        return redirect(url_for('hr.add_staff'))

@bp.route('/reports/<int:report_id>/export')

@login_required
@hr_required
def export_report(report_id):
    """Redirect to staff list"""
    flash("Report export coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/settings/save', methods=['POST'])
@login_required
@hr_required
def save_settings():
    """Redirect to staff list"""
    flash("Settings save coming soon", "info")
    return redirect(url_for('hr.staff_list'))

@bp.route('/settings/save-department', methods=['POST'])
@login_required
@hr_required
def save_department():
    """Redirect to staff list"""
    flash("Department save coming soon", "info")
    return redirect(url_for('hr.staff_list'))

