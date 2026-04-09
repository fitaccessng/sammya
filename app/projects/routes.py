"""
Project Management Module - Complete Routes
Handles all project lifecycle, staff, materials, equipment, DPR, documents, BOQ, analytics.
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import current_user, login_required
from app.models import (
    db, Project, User, BOQItem, PaymentRequest, PaymentRecord, Vendor,
    Milestone, DailyProductionReport, ProjectMaterial, ProjectEquipment,
    ProjectDocument, ProjectBudgetRecord, ProjectPaymentRequest
)
from app.utils import role_required, Roles
from datetime import datetime, timedelta
from functools import wraps
import os
import secrets
from werkzeug.utils import secure_filename
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import json

# Blueprint configuration
bp = Blueprint('project', __name__, url_prefix='/projects')

# Constants
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'png', 'jpg', 'jpeg'}
UPLOAD_FOLDER = 'uploads/projects'

# ======================== HELPER FUNCTIONS ========================

def allowed_file(filename):
    """Check if file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def secure_upload_filename(filename):
    """Generate secure filename with timestamp."""
    ext = filename.rsplit('.', 1)[1].lower()
    name = secure_filename(filename.rsplit('.', 1)[0])
    return f"{secrets.token_hex(8)}_{name}.{ext}"

def check_project_access(user, project_id):
    """Check if user has access to a specific project."""
    if user.role in ['super_hq', 'admin']:
        return True
    
    from app.models import ProjectStaff
    
    # Check if user is the project manager
    project = Project.query.get(project_id)
    if project and project.project_manager_id == user.id:
        return True
    
    # Check if user is assigned to project via ProjectStaff
    staff_assignment = ProjectStaff.query.filter_by(
        user_id=user.id,
        project_id=project_id,
        is_active=True
    ).first()
    
    if staff_assignment:
        return True
    
    if user.role == 'finance':
        return True  # Finance can see all projects
    
    return False

def get_user_accessible_projects(user):
    """Get list of projects accessible to user based on role."""
    if user.role in ['super_hq', 'admin']:
        return Project.query.all()
    elif user.role == 'project_manager':
        # Get projects where user is project manager OR assigned as PROJECT_MANAGER via ProjectStaff
        from app.models import ProjectStaff
        pm_projects = db.session.query(Project).filter(
            Project.project_manager_id == user.id
        ).all()
        
        # Also get projects where user is assigned via ProjectStaff
        staff_projects = db.session.query(Project).join(
            ProjectStaff
        ).filter(
            ProjectStaff.user_id == user.id,
            ProjectStaff.is_active == True
        ).all()
        
        # Combine and deduplicate
        all_projects = {p.id: p for p in pm_projects + staff_projects}
        return list(all_projects.values())
    elif user.role == 'project_staff':
        # Get projects where user is assigned as staff via ProjectStaff
        from app.models import ProjectStaff
        return db.session.query(Project).join(
            ProjectStaff
        ).filter(
            ProjectStaff.user_id == user.id,
            ProjectStaff.is_active == True
        ).all()
    elif user.role == 'finance':
        return Project.query.all()  # Finance can see all for budget tracking
    else:
        return []

def get_user_accessible_project_ids(user):
    """Get list of accessible project IDs for user."""
    projects = get_user_accessible_projects(user)
    return [p.id for p in projects]

def calculate_project_health(project):
    """Calculate project health score (0-100)."""
    health = 100
    
    # Schedule health
    milestones = project.milestones if hasattr(project, 'milestones') else []
    if milestones:
        overdue = sum(1 for m in milestones if m.status == 'delayed')
        health -= min(overdue * 10, 30)
    
    # Budget health
    if project.budget and project.budget > 0:
        expenses = project.expenses if hasattr(project, 'expenses') else []
        spent = sum(float(e.amount or 0) for e in expenses)
        budget_float = float(project.budget)
        if spent > budget_float * 1.1:
            health -= 20
        elif spent > budget_float:
            health -= 10
    
    # Team health
    staff = project.staff_assignments if hasattr(project, 'staff_assignments') else []
    if not staff or len(staff) == 0:
        health -= 15
    
    return max(health, 0)

def calculate_evm_metrics(project):
    """Calculate Earned Value Management metrics."""
    metrics = {
        'planned_value': float(project.budget or 0),
        'earned_value': 0,
        'actual_cost': 0,
        'schedule_variance': 0,
        'cost_variance': 0,
        'schedule_performance_index': 1.0,
        'cost_performance_index': 1.0,
    }
    
    # Calculate earned value from milestones
    milestones = project.milestones if hasattr(project, 'milestones') else []
    total_milestones = len(milestones)
    if total_milestones > 0:
        completed = sum(1 for m in milestones if m.status == 'completed')
        metrics['earned_value'] = (completed / total_milestones) * metrics['planned_value']
    
    # Actual cost from expenses
    expenses = project.expenses if hasattr(project, 'expenses') else []
    metrics['actual_cost'] = sum(float(e.amount or 0) for e in expenses)
    
    # Schedule variance
    if project.start_date:
        # Convert datetime to date if needed for comparison
        start_date = project.start_date.date() if hasattr(project.start_date, 'date') else project.start_date
        current_date = datetime.now().date()
        days_elapsed = (current_date - start_date).days
    else:
        days_elapsed = 0
    
    if project.start_date and project.end_date:
        start_date = project.start_date.date() if hasattr(project.start_date, 'date') else project.start_date
        end_date = project.end_date.date() if hasattr(project.end_date, 'date') else project.end_date
        days_planned = (end_date - start_date).days
    else:
        days_planned = 0
    
    metrics['schedule_variance'] = metrics['earned_value'] - (days_elapsed / max(days_planned, 1)) * metrics['planned_value']
    
    # Cost variance
    metrics['cost_variance'] = metrics['earned_value'] - metrics['actual_cost']
    
    # Performance indices
    if metrics['planned_value'] > 0:
        metrics['schedule_performance_index'] = metrics['earned_value'] / metrics['planned_value']
    if metrics['actual_cost'] > 0:
        metrics['cost_performance_index'] = metrics['earned_value'] / metrics['actual_cost']
    
    return metrics

# ======================== PROJECT DASHBOARD & HOME ========================

@bp.route('/')
@bp.route('/index')
@login_required
def index():
    """Main projects dashboard."""
    accessible_projects = get_user_accessible_projects(current_user)
    
    # Project statistics
    project_count = len(accessible_projects)
    active_projects = sum(1 for p in accessible_projects if p.status == 'active')
    completed_projects = sum(1 for p in accessible_projects if p.status == 'completed')
    total_budget = sum(float(p.budget or 0) for p in accessible_projects)
    completion_rate = (completed_projects / project_count * 100) if project_count > 0 else 0
    
    # Recent projects
    recent_projects = sorted(
        accessible_projects,
        key=lambda p: p.created_at if hasattr(p, 'created_at') else datetime.now(),
        reverse=True
    )[:5]
    
    return render_template(
        'project_manager/index.html',
        accessible_projects=accessible_projects,
        projects=accessible_projects,
        project_stats={
            'total_projects': project_count,
            'active_projects': active_projects,
            'completed_projects': completed_projects,
            'total_budget': total_budget,
            'completion_rate': completion_rate
        },
        recent_projects=recent_projects
    )


@bp.route('/home')
@login_required
def home():
    """Project home page."""
    return redirect(url_for('projects.index'))


@bp.route('/dashboard')
@login_required
def dashboard():
    """Project manager dashboard - overview of projects."""
    accessible_projects = get_user_accessible_projects(current_user)
    
    # Project statistics
    project_count = len(accessible_projects)
    active_projects = sum(1 for p in accessible_projects if p.status == 'active')
    completed_projects = sum(1 for p in accessible_projects if p.status == 'completed')
    total_budget = sum(float(p.budget or 0) for p in accessible_projects)
    completion_rate = (completed_projects / project_count * 100) if project_count > 0 else 0
    
    # Recent projects
    recent_projects = sorted(
        accessible_projects,
        key=lambda p: p.created_at if hasattr(p, 'created_at') else datetime.now(),
        reverse=True
    )[:5]
    
    # Calculate budget utilization
    spent_budget = 0
    for project in accessible_projects:
        if hasattr(project, 'spent') and project.spent:
            spent_budget += float(project.spent)
    
    budget_utilization = 0
    if total_budget > 0:
        budget_utilization = round((spent_budget / total_budget) * 100, 1)
    
    stats = {
        'total': project_count,
        'active': active_projects,
        'completed': completed_projects,
        'budget': total_budget,
        'spent': spent_budget,
        'budget_utilization': budget_utilization
    }
    
    return render_template(
        'project_manager/dashboard.html',
        accessible_projects=accessible_projects,
        projects=accessible_projects,
        stats=stats,
        total_projects=project_count,
        active_projects=active_projects,
        completed_projects=completed_projects,
        total_budget=total_budget,
        completion_rate=completion_rate,
        recent_projects=recent_projects
    )


@bp.route('/staff-dashboard')
@login_required
@role_required(['super_hq', 'project_staff'])
def staff_dashboard():
    """Project staff dashboard for PM-shared execution data."""
    accessible_projects = get_user_accessible_projects(current_user)
    project_id = request.args.get('project_id', type=int)

    if not accessible_projects:
        return render_template(
            'project_staff/dashboard.html',
            accessible_projects=[],
            projects=[],
            project=None,
            dprs=[],
            boq_items=[],
            materials=[],
            milestones=[],
            stats={
                'dpr_count': 0,
                'boq_count': 0,
                'material_count': 0,
                'milestone_count': 0
            }
        )

    project_lookup = {p.id: p for p in accessible_projects}
    if not project_id or project_id not in project_lookup:
        project_id = accessible_projects[0].id

    project = project_lookup[project_id]

    dprs = DailyProductionReport.query.filter_by(project_id=project.id).order_by(
        DailyProductionReport.report_date.desc()
    ).limit(15).all()
    boq_items = BOQItem.query.filter_by(project_id=project.id).order_by(
        BOQItem.updated_at.desc()
    ).limit(20).all()
    materials = ProjectMaterial.query.filter_by(project_id=project.id).order_by(
        ProjectMaterial.created_at.desc()
    ).limit(20).all()
    milestones = Milestone.query.filter_by(project_id=project.id).order_by(
        Milestone.planned_end_date.asc()
    ).all()

    stats = {
        'dpr_count': len(dprs),
        'boq_count': BOQItem.query.filter_by(project_id=project.id).count(),
        'material_count': ProjectMaterial.query.filter_by(project_id=project.id).count(),
        'milestone_count': len(milestones)
    }

    return render_template(
        'project_staff/dashboard.html',
        accessible_projects=accessible_projects,
        projects=accessible_projects,
        project=project,
        dprs=dprs,
        boq_items=boq_items,
        materials=materials,
        milestones=milestones,
        stats=stats
    )


@bp.route('/profile')
@login_required
@role_required(['super_hq', 'project_staff'])
def profile():
    """Project staff profile page."""
    accessible_projects = get_user_accessible_projects(current_user)
    project_id = request.args.get('project_id', type=int)
    selected_project = None

    if accessible_projects:
        project_lookup = {p.id: p for p in accessible_projects}
        if project_id and project_id in project_lookup:
            selected_project = project_lookup[project_id]
        else:
            selected_project = accessible_projects[0]

    return render_template(
        'project_staff/profile.html',
        accessible_projects=accessible_projects,
        projects=accessible_projects,
        project=selected_project
    )

# ======================== PROJECT CRUD ========================

@bp.route('/create', methods=['GET', 'POST'])
@login_required
@role_required(['super_hq', 'project_manager'])
def create_project():
    """Create new project."""
    if request.method == 'POST':
        try:
            project = Project(
                name=request.form.get('name'),
                description=request.form.get('description'),
                start_date=datetime.strptime(request.form.get('start_date'), '%Y-%m-%d'),
                end_date=datetime.strptime(request.form.get('end_date'), '%Y-%m-%d'),
                budget=float(request.form.get('budget') or 0),
                status=request.form.get('status', 'planning'),
                project_manager_id=current_user.id if current_user.role == 'project_manager' else None
            )
            
            db.session.add(project)
            db.session.commit()
            
            flash(f'Project "{project.name}" created successfully', 'success')
            return redirect(url_for('project.project_details', project_id=project.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating project: {str(e)}', 'error')
    
    accessible_projects = get_user_accessible_projects(current_user)
    return render_template('project_manager/create.html', accessible_projects=accessible_projects)

@bp.route('/<int:project_id>')
@login_required
def project_details(project_id):
    """View project details."""
    project = Project.query.get_or_404(project_id)
    
    # Check access
    accessible_ids = get_user_accessible_project_ids(current_user)
    if project_id not in accessible_ids and current_user.role not in ['super_hq', 'admin']:
        flash('Access denied', 'error')
        return redirect(url_for('project.index'))
    
    # Get related data
    staff = project.staff_assignments if hasattr(project, 'staff_assignments') else []
    milestones = project.milestones if hasattr(project, 'milestones') else []
    budget_data = project.budget_records if hasattr(project, 'budget_records') else []
    documents = project.documents if hasattr(project, 'documents') else []
    
    # Calculate metrics
    health = calculate_project_health(project)
    evm = calculate_evm_metrics(project)
    
    return render_template(
        'project_manager/project_details.html',
        project=project,
        staff=staff,
        milestones=milestones,
        budget_data=budget_data,
        documents=documents,
        health=health,
        evm=evm,
        accessible_projects=get_user_accessible_projects(current_user)
    )

@bp.route('/<int:project_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required(['super_hq', 'project_manager'])
def edit_project(project_id):
    """Edit project."""
    project = Project.query.get_or_404(project_id)
    
    # Check project-level access (project manager can only edit their own projects)
    if current_user.role == 'project_manager' and project.project_manager_id != current_user.id:
        flash('You can only edit your own projects', 'error')
        return redirect(url_for('project.dashboard'))
    
    if request.method == 'POST':
        try:
            project.name = request.form.get('name')
            project.description = request.form.get('description')
            project.status = request.form.get('status')
            
            if request.form.get('start_date'):
                project.start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d')
            if request.form.get('end_date'):
                project.end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d')
            if request.form.get('budget'):
                project.budget = float(request.form.get('budget'))
            
            db.session.commit()
            flash('Project updated successfully', 'success')
            return redirect(url_for('project.project_details', project_id=project.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating project: {str(e)}', 'error')
    
    return render_template(
        'project_manager/edit.html',
        project=project,
        accessible_projects=get_user_accessible_projects(current_user)
    )

@bp.route('/<int:project_id>/delete', methods=['POST'])
@login_required
@role_required(['super_hq', 'project_manager'])
def delete_project(project_id):
    """Delete project."""
    project = Project.query.get_or_404(project_id)
    
    try:
        db.session.delete(project)
        db.session.commit()
        flash('Project deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting project: {str(e)}', 'error')
    
    return redirect(url_for('project.index'))

# ======================== STAFF MANAGEMENT ========================

@bp.route('/<int:project_id>/staff')
@login_required
def staff_index(project_id):
    """List project staff."""
    from app.models import ProjectStaff
    
    project = Project.query.get_or_404(project_id)
    
    # Check access
    accessible_ids = get_user_accessible_project_ids(current_user)
    if project_id not in accessible_ids and current_user.role not in ['super_hq', 'admin']:
        return jsonify({'error': 'Access denied'}), 403
    
    role_filter = request.args.get('role_filter', '')
    
    # Get all staff assignments for this project
    staff_query = ProjectStaff.query.filter_by(project_id=project_id, is_active=True)
    
    if role_filter:
        staff_query = staff_query.filter_by(role=role_filter)
    
    all_staff = staff_query.all()
    
    # Categorize by role
    managers = [s for s in all_staff if s.role in ['project_manager', 'Site Engineer', 'site_engineer']]
    supervisors = [s for s in all_staff if s.role in ['SUPERVISOR', 'Supervisor']]
    laborers = [s for s in all_staff if s.role in ['LABORER', 'Laborer', 'Worker']]
    
    staff_stats = {
        'total': len(all_staff),
        'active': sum(1 for s in all_staff if s.is_active),
        'managers': len(managers),
        'field_staff': len(supervisors) + len(laborers)
    }
    
    # Get available users (not yet assigned to this project)
    assigned_user_ids = [s.user_id for s in all_staff]
    available_users = User.query.filter(
        ~User.id.in_(assigned_user_ids) if assigned_user_ids else True
    ).all() if assigned_user_ids else User.query.all()
    
    return render_template(
        'project_manager/staff.html',
        project=project,
        project_id=project_id,
        managers=managers,
        supervisors=supervisors,
        laborers=laborers,
        staff_stats=staff_stats,
        available_employees=available_users,
        accessible_projects=get_user_accessible_projects(current_user),
        selected_project=project
    )

@bp.route('/<int:project_id>/staff/assign', methods=['POST'])
@login_required
@role_required(['super_hq', 'project_manager', 'admin'])
def assign_staff(project_id):
    """Assign staff to project."""
    from app.models import ProjectStaff
    from datetime import datetime as dt
    
    project = Project.query.get_or_404(project_id)
    
    try:
        user_id = request.form.get('employee_id') or request.form.get('user_id')
        role = request.form.get('role', 'Project Staff')
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        
        # Check if already assigned
        existing = ProjectStaff.query.filter_by(user_id=user_id, project_id=project_id).first()
        
        if existing:
            existing.role = role
            existing.is_active = True
            if start_date_str:
                existing.start_date = dt.strptime(start_date_str, '%Y-%m-%d').date()
            if end_date_str:
                existing.end_date = dt.strptime(end_date_str, '%Y-%m-%d').date()
            flash('Staff role updated successfully', 'success')
        else:
            assignment = ProjectStaff(
                user_id=user_id,
                project_id=project_id,
                role=role,
                is_active=True
            )
            if start_date_str:
                assignment.start_date = dt.strptime(start_date_str, '%Y-%m-%d').date()
            if end_date_str:
                assignment.end_date = dt.strptime(end_date_str, '%Y-%m-%d').date()
            
            db.session.add(assignment)
            flash('Staff assigned successfully', 'success')
        
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Error assigning staff: {str(e)}', 'error')
    
    return redirect(url_for('project.staff_index', project_id=project_id))

@bp.route('/<int:project_id>/staff/<int:staff_id>/remove', methods=['DELETE'])
@login_required
@role_required(['super_hq', 'project_manager', 'admin'])
def remove_staff(project_id, staff_id):
    """Remove staff from project."""
    from app.models import ProjectStaff
    
    assignment = ProjectStaff.query.get_or_404(staff_id)
    
    try:
        db.session.delete(assignment)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400

# ======================== MATERIALS MANAGEMENT ========================

@bp.route('/<int:project_id>/materials')
@login_required
def materials_index(project_id):
    """List project materials."""
    project = Project.query.get_or_404(project_id)
    
    # Check access
    accessible_ids = get_user_accessible_project_ids(current_user)
    if project_id not in accessible_ids and current_user.role not in ['super_hq', 'ADMIN']:
        return jsonify({'error': 'Access denied'}), 403
    
    category_filter = request.args.get('category', '')
    materials = project.materials if hasattr(project, 'materials') else []
    
    if category_filter:
        materials = [m for m in materials if m.category == category_filter]
    
    material_stats = {
        'total': len(project.materials if hasattr(project, 'materials') else []),
        'in_stock': sum(1 for m in (project.materials if hasattr(project, 'materials') else []) if m.quantity_allocated > (m.quantity_used or 0)),
        'used': sum(float(m.quantity_used or 0) for m in (project.materials if hasattr(project, 'materials') else [])),
        'total_cost': sum(float(m.unit_cost * m.quantity_allocated) for m in (project.materials if hasattr(project, 'materials') else []))
    }
    
    return render_template(
        'project_manager/materials.html',
        project=project,
        project_materials=materials,
        material_stats=material_stats,
        accessible_projects=get_user_accessible_projects(current_user),
        selected_project=project
    )

@bp.route('/<int:project_id>/materials/add', methods=['POST'])
@login_required
@role_required(['super_hq', 'project_manager', 'project_staff'])
def add_material(project_id):
    """Add material to project."""
    project = Project.query.get_or_404(project_id)
    
    try:
        material = Material(
            project_id=project_id,
            description=request.form.get('description'),
            category=request.form.get('category'),
            unit=request.form.get('unit'),
            quantity_allocated=float(request.form.get('quantity')),
            unit_cost=float(request.form.get('unit_cost')),
            specification=request.form.get('specification')
        )
        
        db.session.add(material)
        db.session.commit()
        
        flash('Material added successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding material: {str(e)}', 'error')
    
    return redirect(url_for('project.materials_index', project_id=project_id))

@bp.route('/<int:project_id>/materials/<int:material_id>/delete', methods=['DELETE'])
@login_required
@role_required(['super_hq', 'project_manager'])
def delete_material(project_id, material_id):
    """Delete material."""
    material = Material.query.get_or_404(material_id)
    
    try:
        db.session.delete(material)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400

# ======================== EQUIPMENT MANAGEMENT ========================

@bp.route('/<int:project_id>/equipment')
@login_required
def equipment_index(project_id):
    """List project equipment."""
    project = Project.query.get_or_404(project_id)
    
    # Check access
    accessible_ids = get_user_accessible_project_ids(current_user)
    if project_id not in accessible_ids and current_user.role not in ['super_hq', 'ADMIN']:
        return jsonify({'error': 'Access denied'}), 403
    
    equipment = project.equipment if hasattr(project, 'equipment') else []
    
    return render_template(
        'project_manager/equipment.html',
        project=project,
        equipment=equipment,
        accessible_projects=get_user_accessible_projects(current_user),
        selected_project=project
    )

@bp.route('/<int:project_id>/equipment/add', methods=['POST'])
@login_required
@role_required(['super_hq', 'project_manager', 'project_staff'])
def add_equipment(project_id):
    """Add equipment to project."""
    project = Project.query.get_or_404(project_id)
    
    try:
        equipment = Equipment(
            project_id=project_id,
            name=request.form.get('name'),
            type=request.form.get('type'),
            description=request.form.get('description'),
            status='operational'
        )
        
        db.session.add(equipment)
        db.session.commit()
        
        flash('Equipment added successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding equipment: {str(e)}', 'error')
    
    return redirect(url_for('project.equipment_index', project_id=project_id))

# ======================== TIMELINE & MILESTONES ========================

@bp.route('/<int:project_id>/timeline')
@login_required
def timeline_index(project_id):
    """View project timeline and milestones."""
    project = Project.query.get_or_404(project_id)
    
    # Check access
    accessible_ids = get_user_accessible_project_ids(current_user)
    if project_id not in accessible_ids and current_user.role not in ['super_hq', 'admin']:
        return jsonify({'error': 'Access denied'}), 403
    
    status_filter = request.args.get('status', '')
    milestones = project.milestones if hasattr(project, 'milestones') else []
    
    if status_filter:
        milestones = [m for m in milestones if m.status == status_filter]
    
    milestone_stats = {
        'total': len(project.milestones if hasattr(project, 'milestones') else []),
        'completed': sum(1 for m in (project.milestones if hasattr(project, 'milestones') else []) if m.status == 'completed'),
        'in_progress': sum(1 for m in (project.milestones if hasattr(project, 'milestones') else []) if m.status == 'in_progress'),
        'completion_percentage': 0
    }
    
    if milestone_stats['total'] > 0:
        milestone_stats['completion_percentage'] = int(
            (milestone_stats['completed'] / milestone_stats['total']) * 100
        )
    
    return render_template(
        'project_manager/timeline.html',
        project=project,
        project_milestones=milestones,
        milestone_stats=milestone_stats,
        accessible_projects=get_user_accessible_projects(current_user),
        selected_project=project
    )

@bp.route('/<int:project_id>/milestones/add', methods=['POST'])
@login_required
@role_required(['super_hq', 'project_manager', 'project_staff'])
def add_milestone(project_id):
    """Add milestone to project."""
    project = Project.query.get_or_404(project_id)
    
    # Check project-level access
    if not check_project_access(current_user, project_id):
        flash('You do not have access to this project', 'error')
        return redirect(url_for('project.dashboard'))
    
    try:
        milestone = Milestone(
            project_id=project_id,
            name=request.form.get('name'),
            description=request.form.get('description'),
            planned_start_date=datetime.strptime(request.form.get('start_date'), '%Y-%m-%d'),
            planned_end_date=datetime.strptime(request.form.get('end_date'), '%Y-%m-%d'),
            deliverables=request.form.get('deliverables'),
            status=request.form.get('status', 'not_started'),
            completion_percentage=0
        )
        
        db.session.add(milestone)
        db.session.commit()
        
        flash('Milestone added successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding milestone: {str(e)}', 'error')
    
    return redirect(url_for('project.timeline_index', project_id=project_id))

@bp.route('/<int:project_id>/milestones/<int:milestone_id>/delete', methods=['DELETE'])
@login_required
@role_required(['super_hq', 'project_manager'])
def delete_milestone(project_id, milestone_id):
    """Delete milestone."""
    milestone = Milestone.query.get_or_404(milestone_id)
    
    try:
        db.session.delete(milestone)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400

# ======================== DAILY PRODUCTION REPORTS (DPR) ========================

@bp.route('/<int:project_id>/dpr')
@login_required
def dpr_index(project_id):
    """List Daily Production Reports."""
    project = Project.query.get_or_404(project_id)
    
    # Check access
    accessible_ids = get_user_accessible_project_ids(current_user)
    if project_id not in accessible_ids and current_user.role not in ['super_hq', 'admin']:
        return jsonify({'error': 'Access denied'}), 403
    
    dprs = project.daily_reports if hasattr(project, 'daily_reports') else []
    
    dpr_summary = {
        'total': len(dprs),
        'completed': sum(1 for d in dprs if d.status == 'completed'),
        'pending': sum(1 for d in dprs if d.status == 'sent_to_staff'),
        'monthly': sum(1 for d in dprs if d.report_date and d.report_date.month == datetime.now().month)
    }
    
    return render_template(
        'project_manager/dpr.html',
        project=project,
        project_dprs=dprs,
        dpr_summary=dpr_summary,
        accessible_projects=get_user_accessible_projects(current_user),
        selected_project=project
    )

@bp.route('/<int:project_id>/dpr/create', methods=['POST'])
@login_required
@role_required(['super_hq', 'project_manager', 'project_staff', 'site_engineer'])
def create_dpr(project_id):
    """Create new DPR."""
    project = Project.query.get_or_404(project_id)
    
    # Check project-level access
    if not check_project_access(current_user, project_id):
        flash('You do not have access to this project', 'error')
        return redirect(url_for('project.dashboard'))
    
    try:
        dpr = DailyProductionReport(
            project_id=project_id,
            report_date=datetime.strptime(request.form.get('report_date'), '%Y-%m-%d'),
            status='draft',
            created_by=getattr(current_user, 'username', current_user.name),
            weather_conditions=request.form.get('weather', ''),
            work_description=request.form.get('work_description', ''),
            unit=request.form.get('unit', ''),
            staff_report=request.form.get('staff_report', ''),
            general_remarks=request.form.get('remarks', '')
        )
        
        db.session.add(dpr)
        db.session.commit()
        
        flash('DPR created successfully', 'success')
        return redirect(url_for('project.view_dpr', dpr_id=dpr.id))
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating DPR: {str(e)}', 'error')
        return redirect(url_for('project.dpr_index', project_id=project_id))

@bp.route('/dpr/<int:dpr_id>')
@login_required
def view_dpr(dpr_id):
    """View DPR details."""
    dpr = DailyProductionReport.query.get_or_404(dpr_id)
    
    # Check access
    accessible_ids = get_user_accessible_project_ids(current_user)
    if dpr.project_id not in accessible_ids and current_user.role not in ['super_hq', 'ADMIN']:
        flash('Access denied', 'error')
        return redirect(url_for('project.index'))
    
    return render_template(
        'project_manager/dpr_detail.html',
        dpr=dpr,
        project=dpr.project,
        accessible_projects=get_user_accessible_projects(current_user)
    )

@bp.route('/dpr/<int:dpr_id>/submit', methods=['POST'])
@login_required
def submit_dpr(dpr_id):
    """Submit DPR for approval."""
    dpr = DailyProductionReport.query.get_or_404(dpr_id)
    
    # Check project-level access
    if not check_project_access(current_user, dpr.project_id):
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    try:
        dpr.status = 'sent_to_staff'
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400

@bp.route('/dpr/<int:dpr_id>/approve', methods=['POST'])
@login_required
@role_required(['super_hq', 'project_manager'])
def approve_dpr(dpr_id):
    """Approve DPR."""
    dpr = DailyProductionReport.query.get_or_404(dpr_id)
    
    # Check project-level access
    if not check_project_access(current_user, dpr.project_id):
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    try:
        dpr.status = 'completed'
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400

@bp.route('/dpr/<int:dpr_id>/reject', methods=['POST'])
@login_required
@role_required(['super_hq', 'project_manager'])
def reject_dpr(dpr_id):
    """Reject DPR."""
    dpr = DailyProductionReport.query.get_or_404(dpr_id)
    
    # Check project-level access
    if not check_project_access(current_user, dpr.project_id):
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    data = request.get_json()
    
    try:
        dpr.status = 'rejected'
        if data.get('reason'):
            dpr.rejection_reason = data.get('reason')
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400

# ======================== DOCUMENTS MANAGEMENT ========================

@bp.route('/<int:project_id>/documents')
@login_required
def documents_index(project_id):
    """List project documents."""
    project = Project.query.get_or_404(project_id)
    
    # Check access
    accessible_ids = get_user_accessible_project_ids(current_user)
    if project_id not in accessible_ids and current_user.role not in ['super_hq', 'ADMIN']:
        return jsonify({'error': 'Access denied'}), 403
    
    doc_type = request.args.get('doc_type', '')
    approval_status = request.args.get('approval_status', '')
    
    documents = project.documents if hasattr(project, 'documents') else []
    
    if doc_type:
        documents = [d for d in documents if d.document_type == doc_type]
    if approval_status:
        documents = [d for d in documents if d.approval_state == approval_status]
    
    doc_stats = {
        'total': len(project.documents if hasattr(project, 'documents') else []),
        'pending': sum(1 for d in (project.documents if hasattr(project, 'documents') else []) if d.approval_state == 'pending'),
        'approved': sum(1 for d in (project.documents if hasattr(project, 'documents') else []) if d.approval_state == 'approved'),
        'rejected': sum(1 for d in (project.documents if hasattr(project, 'documents') else []) if d.approval_state == 'rejected')
    }
    
    return render_template(
        'project_manager/documents.html',
        project=project,
        documents=documents,
        doc_stats=doc_stats,
        accessible_projects=get_user_accessible_projects(current_user),
        selected_project=project
    )

@bp.route('/<int:project_id>/documents/upload', methods=['POST'])
@login_required
@role_required(['super_hq', 'project_manager', 'project_staff'])
def upload_document(project_id):
    """Upload project document."""
    project = Project.query.get_or_404(project_id)
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'File type not allowed'}), 400
    
    try:
        filename = secure_upload_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        file.save(filepath)
        
        document = Document(
            project_id=project_id,
            title=request.form.get('title'),
            document_type=request.form.get('doc_type'),
            description=request.form.get('description'),
            file_path=filepath,
            uploaded_by=current_user.username,
            upload_date=datetime.now(),
            approval_state='pending'
        )
        
        db.session.add(document)
        db.session.commit()
        
        flash('Document uploaded successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error uploading document: {str(e)}', 'error')
    
    return redirect(url_for('project.documents_index', project_id=project_id))

@bp.route('/documents/<int:doc_id>/download')
@login_required
def download_document(doc_id):
    """Download document."""
    document = Document.query.get_or_404(doc_id)
    
    # Check access
    accessible_ids = get_user_accessible_project_ids(current_user)
    if document.project_id not in accessible_ids and current_user.role not in ['super_hq', 'ADMIN']:
        flash('Access denied', 'error')
        return redirect(url_for('project.index'))
    
    try:
        return send_file(document.file_path, as_attachment=True)
    except Exception as e:
        flash(f'Error downloading document: {str(e)}', 'error')
        return redirect(url_for('project.documents_index', project_id=document.project_id))

@bp.route('/documents/<int:doc_id>/approve', methods=['POST'])
@login_required
@role_required(['super_hq'])
def approve_document(doc_id):
    """Approve document."""
    document = Document.query.get_or_404(doc_id)
    
    try:
        document.approval_state = 'approved'
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400

@bp.route('/documents/<int:doc_id>/reject', methods=['POST'])
@login_required
@role_required(['super_hq'])
def reject_document(doc_id):
    """Reject document."""
    document = Document.query.get_or_404(doc_id)
    data = request.get_json()
    
    try:
        document.approval_state = 'rejected'
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400

# ======================== REPORTS ========================

@bp.route('/<int:project_id>/reports')
@login_required
def reports_index(project_id):
    """List project reports."""
    project = Project.query.get_or_404(project_id)
    
    # Check access
    accessible_ids = get_user_accessible_project_ids(current_user)
    if project_id not in accessible_ids and current_user.role not in ['super_hq', 'ADMIN']:
        return jsonify({'error': 'Access denied'}), 403
    
    dprs = project.daily_reports if hasattr(project, 'daily_reports') else []
    reports = project.reports if hasattr(project, 'reports') else []
    
    dpr_summary = {
        'total': len(dprs),
        'completed': sum(1 for d in dprs if d.status == 'completed'),
        'pending': sum(1 for d in dprs if d.status == 'sent_to_staff'),
        'monthly': sum(1 for d in dprs if d.report_date and d.report_date.month == datetime.now().month)
    }
    
    return render_template(
        'projects/reports.html',
        project=project,
        project_dprs=dprs,
        project_reports=reports,
        dpr_summary=dpr_summary,
        accessible_projects=get_user_accessible_projects(current_user),
        selected_project=project
    )

@bp.route('/reports/<int:report_id>')
@login_required
def view_report(report_id):
    """View report details."""
    report = Report.query.get_or_404(report_id)
    
    # Check access
    accessible_ids = get_user_accessible_project_ids(current_user)
    if report.project_id not in accessible_ids and current_user.role not in ['super_hq', 'ADMIN']:
        flash('Access denied', 'error')
        return redirect(url_for('project.index'))
    
    return render_template('project_manager/reports.html')


@bp.route('/<int:project_id>/payment-request/create', methods=['POST'])
@login_required
@role_required(['super_hq', 'project_manager', 'project_staff'])
def create_project_payment_request(project_id):
    """Create project-specific payment request for finance approval."""
    project = Project.query.get_or_404(project_id)
    if not check_project_access(current_user, project_id):
        flash('Access denied', 'error')
        return redirect(url_for('project.project_details', project_id=project_id))

    try:
        title = request.form.get('title', '').strip()
        amount = float(request.form.get('amount') or 0)
        description = request.form.get('description', '').strip()
        if not title or amount <= 0:
            flash('Title and valid amount are required', 'error')
            return redirect(url_for('project.project_details', project_id=project_id))

        req = ProjectPaymentRequest(
            project_id=project.id,
            requested_by=current_user.id,
            title=title,
            description=description,
            amount=amount,
            approval_state='pending'
        )
        db.session.add(req)
        db.session.commit()
        flash('Payment request submitted to Finance for approval.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating payment request: {str(e)}', 'error')

    return redirect(url_for('project.project_details', project_id=project_id))

# ======================== ANALYTICS ========================

@bp.route('/<int:project_id>/analytics')
@login_required
def analytics_index(project_id):
    """View project analytics."""
    project = Project.query.get_or_404(project_id)
    
    # Check access
    accessible_ids = get_user_accessible_project_ids(current_user)
    if project_id not in accessible_ids and current_user.role not in ['super_hq', 'ADMIN', 'FINANCE']:
        return jsonify({'error': 'Access denied'}), 403
    
    evm = calculate_evm_metrics(project)
    health_score = calculate_project_health(project)
    
    # Budget breakdown
    evm['labor_cost'] = 0
    evm['material_cost'] = sum(float(m.unit_cost * m.quantity_allocated) for m in (project.materials if hasattr(project, 'materials') else []))
    evm['equipment_cost'] = 0
    evm['other_cost'] = evm['actual_cost'] - evm['material_cost']
    evm['total_budget'] = float(project.budget or 0)
    evm['spent_amount'] = float(evm['actual_cost'] or 0)
    evm['remaining_amount'] = max(0.0, evm['total_budget'] - evm['spent_amount'])
    evm['budget_utilization'] = round((evm['spent_amount'] / evm['total_budget']) * 100, 1) if evm['total_budget'] > 0 else 0
    evm['budget_spent'] = evm['spent_amount']
    evm['budget_remaining'] = evm['remaining_amount']
    evm['overall_progress'] = int(min(100, (evm['earned_value'] / max(evm['planned_value'], 1)) * 100))
    evm['schedule_performance'] = round(evm['schedule_performance_index'] * 100, 1)
    evm['cost_performance'] = round(evm['cost_performance_index'] * 100, 1)
    
    # Staff metrics
    staff = project.staff_assignments if hasattr(project, 'staff_assignments') else []
    evm['total_staff_count'] = len(staff)
    evm['staff_by_role'] = {
        'managers': sum(1 for s in staff if s.role in ['project_manager', 'site_engineer']),
        'engineers': sum(1 for s in staff if s.role == 'site_engineer'),
        'supervisors': sum(1 for s in staff if s.role == 'SUPERVISOR'),
        'laborers': sum(1 for s in staff if s.role == 'LABORER')
    }
    evm['total_team_members'] = evm['total_staff_count']
    
    # Equipment metrics
    equipment = project.equipment if hasattr(project, 'equipment') else []
    evm['total_equipment'] = len(equipment)
    evm['active_equipment'] = sum(1 for e in equipment if e.status == 'operational')
    evm['total_equipment_hours'] = sum(float(getattr(e, 'hours_used', 0)) for e in equipment)
    evm['equipment_count'] = len(equipment)
    evm['materials_count'] = len(project.materials if hasattr(project, 'materials') else [])
    evm['document_count'] = len(project.documents if hasattr(project, 'documents') else [])

    milestones = project.milestones if hasattr(project, 'milestones') else []
    evm['total_milestones'] = len(milestones)
    evm['completed_milestones'] = sum(1 for m in milestones if m.status == 'completed')
    evm['in_progress_milestones'] = sum(1 for m in milestones if m.status == 'in_progress')
    evm['delayed_milestones'] = sum(1 for m in milestones if m.status == 'delayed')
    evm['completion_percentage'] = round((evm['completed_milestones'] / evm['total_milestones']) * 100, 1) if evm['total_milestones'] > 0 else 0
    
    return render_template(
        'project_manager/analytics.html',
        project=project,
        evm=evm,
        health={'overall_health': health_score},
        accessible_projects=get_user_accessible_projects(current_user),
        selected_project=project
    )

# ======================== BOQ (BILL OF QUANTITIES) ========================

@bp.route('/<int:project_id>/boq')
@login_required
def boq_index(project_id):
    """View Bill of Quantities."""
    project = Project.query.get_or_404(project_id)
    
    # Check access
    accessible_ids = get_user_accessible_project_ids(current_user)
    if project_id not in accessible_ids and current_user.role not in ['super_hq', 'ADMIN', 'project_manager']:
        return jsonify({'error': 'Access denied'}), 403
    
    status_filter = request.args.get('status', '')
    boq_items = project.boq_items if hasattr(project, 'boq_items') else []
    
    if status_filter:
        boq_items = [b for b in boq_items if str(getattr(b, 'approval_state', '')).lower() == status_filter.lower()]
    
    boq_summary = {
        'total_items': len(boq_items),
        'total_cost': sum(float(b.amount or 0) for b in boq_items),
        'sourced_items': sum(1 for b in boq_items if str(getattr(b, 'approval_state', '')) == 'approved'),
        'pending_items': sum(1 for b in boq_items if str(getattr(b, 'approval_state', '')) in ['draft', 'pending'])
    }
    
    return render_template(
        'project_manager/boq.html',
        project=project,
        boq_items=boq_items,
        boq_summary=boq_summary,
        accessible_projects=get_user_accessible_projects(current_user),
        selected_project=project
    )

@bp.route('/<int:project_id>/boq/add', methods=['POST'])
@login_required
@role_required(['super_hq', 'project_manager', 'project_staff'])
def add_boq_item(project_id):
    """Add BOQ item."""
    project = Project.query.get_or_404(project_id)
    
    try:
        item = BOQItem(
            project_id=project_id,
            description=(request.form.get('description') or '').strip() or f"BOQ Item {BOQItem.query.filter_by(project_id=project_id).count() + 1}",
            unit=request.form.get('unit'),
            quantity=float(request.form.get('quantity')),
            unit_rate=float(request.form.get('unit_cost') or request.form.get('unit_rate') or 0),
            created_by=current_user.id
        )
        item.calculate_amount()
        db.session.add(item)
        db.session.commit()
        
        flash('BOQ item added successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding BOQ item: {str(e)}', 'error')
    
    return redirect(url_for('project.boq_index', project_id=project_id))

@bp.route('/<int:project_id>/boq/import', methods=['POST'])
@login_required
@role_required(['super_hq', 'project_manager', 'project_staff'])
def import_boq(project_id):
    """Import BOQ from Excel file."""
    project = Project.query.get_or_404(project_id)
    
    file = request.files.get('boq_file') or request.files.get('file')
    if not file:
        flash('No file provided', 'error')
        return redirect(url_for('project.boq_index', project_id=project_id))
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Please upload an Excel file', 'error')
        return redirect(url_for('project.boq_index', project_id=project_id))
    
    try:
        df = pd.read_excel(file)
        
        for idx, row in df.iterrows():
            quantity = float(row.get('Quantity', 0) or 0)
            unit_rate = float(row.get('Unit Rate', 0) or row.get('Unit Price', 0) or 0)
            item = BOQItem(
                project_id=project_id,
                description=str(row.get('Description', '') or f'BOQ Item {idx + 1}'),
                unit=str(row.get('Unit', '')),
                quantity=quantity,
                unit_rate=unit_rate,
                amount=float(row.get('Amount', 0) or (quantity * unit_rate)),
                created_by=current_user.id
            )
            db.session.add(item)
        
        db.session.commit()
        flash(f'Imported {len(df)} BOQ items successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing BOQ: {str(e)}', 'error')
    
    return redirect(url_for('project.boq_index', project_id=project_id))

@bp.route('/<int:project_id>/boq/<int:item_id>/delete', methods=['DELETE'])
@login_required
@role_required(['super_hq', 'project_manager'])
def delete_boq_item(project_id, item_id):
    """Delete BOQ item."""
    item = BOQItem.query.get_or_404(item_id)
    
    try:
        db.session.delete(item)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400

@bp.route('/<int:project_id>/boq/export')
@login_required
def export_boq_excel(project_id):
    """Export BOQ to Excel."""
    project = Project.query.get_or_404(project_id)
    
    # Check access
    accessible_ids = get_user_accessible_project_ids(current_user)
    if project_id not in accessible_ids and current_user.role not in ['super_hq', 'ADMIN', 'project_manager']:
        flash('Access denied', 'error')
        return redirect(url_for('project.index'))
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'BOQ'
        
        # Headers
        headers = ['Description', 'Unit', 'Quantity', 'Unit Rate', 'Amount']
        ws.append(headers)
        
        # Data
        boq_items = project.boq_items if hasattr(project, 'boq_items') else []
        for item in boq_items:
            ws.append([
                item.description,
                item.unit,
                item.quantity,
                item.unit_rate,
                item.amount
            ])
        
        # Save to temp file
        temp_file = f'boq_{project_id}_{datetime.now().timestamp()}.xlsx'
        wb.save(temp_file)
        
        return send_file(temp_file, as_attachment=True, download_name=f'BOQ_{project.name}.xlsx')
    except Exception as e:
        flash(f'Error exporting BOQ: {str(e)}', 'error')
        return redirect(url_for('project.boq_index', project_id=project_id))

@bp.route('/boq/template')
@login_required
def download_boq_template():
    """Download BOQ template."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'BOQ'
        
        headers = ['S/N', 'Description', 'Unit', 'Quantity', 'Unit Rate']
        ws.append(headers)
        
        # Add sample rows
        for i in range(1, 6):
            ws.append([i, f'Item {i}', '', '', ''])
        
        temp_file = f'boq_template_{datetime.now().timestamp()}.xlsx'
        wb.save(temp_file)
        
        return send_file(temp_file, as_attachment=True, download_name='BOQ_Template.xlsx')
    except Exception as e:
        flash(f'Error downloading template: {str(e)}', 'error')
        return redirect(url_for('project.index'))

# ======================== SETTINGS ========================

@bp.route('/settings', methods=['GET', 'POST'])
@login_required
@role_required(['super_hq', 'project_manager'])
def settings():
    """Project settings page."""
    if request.method == 'POST':
        # Update settings logic here
        flash('Settings updated successfully', 'success')
    
    settings_data = {
        'currency': 'NGN',
        'date_format': 'dd-mm-yyyy',
        'enable_approval_workflow': True,
        'default_duration': 90,
        'default_template': 'standard',
        'notify_dpr_created': True,
        'notify_dpr_submitted': True,
        'notify_dpr_approved': True,
        'notify_budget_threshold': True,
        'budget_threshold': 80,
        'notify_milestone_due': True,
        'notify_milestone_overdue': True
    }
    
    return render_template(
        'project_manager/settings.html',
        settings=settings_data,
        now=datetime.now(),
        accessible_projects=get_user_accessible_projects(current_user)
    )

@bp.route('/settings/update', methods=['POST'])
@login_required
@role_required(['super_hq', 'project_manager'])
def update_settings():
    """Update project settings."""
    try:
        # Update settings logic
        flash('Settings updated successfully', 'success')
    except Exception as e:
        flash(f'Error updating settings: {str(e)}', 'error')
    
    return redirect(url_for('project.settings'))

@bp.route('/settings/notifications', methods=['POST'])
@login_required
@role_required(['super_hq', 'project_manager'])
def update_notification_settings():
    """Update notification preferences."""
    try:
        flash('Notification settings updated successfully', 'success')
    except Exception as e:
        flash(f'Error updating settings: {str(e)}', 'error')
    
    return redirect(url_for('project.settings'))

# ======================== EXPORT & REPORTING ========================

@bp.route('/<int:project_id>/analytics/export/pdf')
@login_required
def export_analytics_pdf(project_id):
    """Export analytics as PDF."""
    project = Project.query.get_or_404(project_id)
    
    accessible_ids = get_user_accessible_project_ids(current_user)
    if project_id not in accessible_ids and current_user.role not in ['super_hq', 'ADMIN', 'FINANCE']:
        flash('Access denied', 'error')
        return redirect(url_for('project.index'))
    
    flash('PDF export coming soon', 'info')
    return redirect(url_for('project.analytics_index', project_id=project_id))

@bp.route('/<int:project_id>/analytics/export/excel')
@login_required
def export_analytics_excel(project_id):
    """Export analytics as Excel."""
    project = Project.query.get_or_404(project_id)
    
    accessible_ids = get_user_accessible_project_ids(current_user)
    if project_id not in accessible_ids and current_user.role not in ['super_hq', 'ADMIN', 'FINANCE']:
        flash('Access denied', 'error')
        return redirect(url_for('project.index'))
    
    flash('Excel export coming soon', 'info')
    return redirect(url_for('project.analytics_index', project_id=project_id))

@bp.route('/dpr/<int:dpr_id>/export')
@login_required
def export_dpr(dpr_id):
    """Export DPR as PDF."""
    dpr = DailyProductionReport.query.get_or_404(dpr_id)
    
    accessible_ids = get_user_accessible_project_ids(current_user)
    if dpr.project_id not in accessible_ids and current_user.role not in ['super_hq', 'ADMIN', 'project_manager']:
        flash('Access denied', 'error')
        return redirect(url_for('project.index'))
    
    flash('DPR export coming soon', 'info')
    return redirect(url_for('project.view_dpr', dpr_id=dpr_id))

# ======================== ERROR HANDLERS ========================



# ======================== PROJECT ACTIVITY LOG ========================

@bp.route('/<int:project_id>/activity-log')
@login_required
def activity_log(project_id):
    """View project activity log."""
    from app.models import ProjectActivityLog
    
    project = Project.query.get_or_404(project_id)
    page = request.args.get('page', 1, type=int)
    
    # Get activity logs
    logs = ProjectActivityLog.query.filter_by(project_id=project_id).order_by(
        ProjectActivityLog.timestamp.desc()
    ).paginate(page=page, per_page=20)
    
    return render_template(
        'project_manager/activity_log.html',
        project=project,
        logs=logs
    )


# ======================== ERROR HANDLERS ========================

@bp.errorhandler(404)
def page_not_found(error):
    """Handle 404 errors."""
    return render_template('errors/404.html'), 404

@bp.errorhandler(403)
def access_denied(error):
    """Handle 403 errors."""
    return render_template('errors/403.html'), 403

@bp.errorhandler(500)
def server_error(error):
    """Handle 500 errors."""
    return render_template('errors/500.html'), 500

