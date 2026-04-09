"""
Payroll Export Engines
Generate exports for bank payments, tax remittance, pension, insurance, loans
"""

import csv
import json
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Tuple
from io import StringIO, BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path
from app.payroll_models import PayrollBatch, PayrollRecord, PayrollExport
from app.models import db
import logging
import hashlib

logger = logging.getLogger(__name__)


class PayrollExportEngine:
    """Generate various payroll exports"""
    
    # Export paths
    EXPORT_BASE_PATH = Path('app/uploads/payroll_exports')
    
    @staticmethod
    def generate_bank_payment_export(batch_id: int, export_format: str = 'csv') -> Tuple[bool, Dict]:
        """
        Generate bank bulk payment file
        Format: Bank CSV/TXT for bulk salary transfers
        """
        result = {
            'success': False,
            'file_path': None,
            'file_name': None,
            'record_count': 0,
            'total_amount': Decimal(0),
            'error': None
        }
        
        try:
            batch = PayrollBatch.query.get_or_404(batch_id)
            records = PayrollRecord.query.filter_by(batch_id=batch_id).all()
            
            if not records:
                result['error'] = 'No records to export'
                return False, result
            
            # Create export directory
            PayrollExportEngine.EXPORT_BASE_PATH.mkdir(parents=True, exist_ok=True)
            
            file_name = f"BANK_PAYMENT_{batch.payroll_period.replace('-', '')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{export_format}"
            file_path = PayrollExportEngine.EXPORT_BASE_PATH / file_name
            
            if export_format == 'csv':
                PayrollExportEngine._write_bank_csv(file_path, records)
            elif export_format == 'excel':
                PayrollExportEngine._write_bank_excel(file_path, records)
            elif export_format == 'txt':
                PayrollExportEngine._write_bank_txt(file_path, records)
            
            # Calculate totals
            total_amount = sum(Decimal(r.net_salary or 0) for r in records)
            
            # Create export record
            file_hash = PayrollExportEngine._calculate_file_hash(file_path)
            
            export = PayrollExport(
                batch_id=batch_id,
                export_type='bank_payment',
                export_format=export_format,
                file_name=file_name,
                file_path=str(file_path),
                file_size=file_path.stat().st_size,
                file_hash=file_hash,
                record_count=len(records),
                total_amount=total_amount,
                status='generated'
            )
            db.session.add(export)
            db.session.commit()
            
            result['success'] = True
            result['file_path'] = str(file_path)
            result['file_name'] = file_name
            result['record_count'] = len(records)
            result['total_amount'] = total_amount
            
            return True, result
            
        except Exception as e:
            logger.error(f"Bank export error: {str(e)}")
            result['error'] = str(e)
            return False, result
    
    @staticmethod
    def _write_bank_csv(file_path: Path, records: List[PayrollRecord]):
        """Write bank payment CSV"""
        with open(file_path, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=[
                'Account Number', 'Beneficiary Name', 'Amount', 'Bank Code', 
                'Bank Name', 'Reference', 'Narration'
            ])
            writer.writeheader()
            
            for record in records:
                if record.user and record.net_salary > 0:
                    writer.writerow({
                        'Account Number': record.bank_account or '',
                        'Beneficiary Name': record.user.name,
                        'Amount': float(record.net_salary),
                        'Bank Code': '000',  # Standard code
                        'Bank Name': record.bank_name or 'DEFAULT BANK',
                        'Reference': f'PR{record.id}',
                        'Narration': f'SALARY {record.payroll_period}'
                    })
    
    @staticmethod
    def _write_bank_excel(file_path: Path, records: List[PayrollRecord]):
        """Write bank payment Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Bank Payment'
        
        # Headers
        headers = ['Account Number', 'Beneficiary Name', 'Amount', 'Bank Code', 
                   'Bank Name', 'Reference', 'Narration']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for record in records:
            if record.user and record.net_salary > 0:
                ws.append([
                    record.bank_account or '',
                    record.user.name,
                    float(record.net_salary),
                    '000',
                    record.bank_name or 'DEFAULT BANK',
                    f'PR{record.id}',
                    f'SALARY {record.payroll_period}'
                ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 20
        
        wb.save(file_path)
    
    @staticmethod
    def _write_bank_txt(file_path: Path, records: List[PayrollRecord]):
        """Write bank payment TXT (fixed width format)"""
        with open(file_path, 'w') as f:
            f.write('PAYROLL BANK TRANSFER\n')
            f.write(f'Generated: {datetime.now().isoformat()}\n')
            f.write('-' * 100 + '\n')
            f.write(f'{"ACCOUNT":<20} {"NAME":<30} {"AMOUNT":>15} {"BANK":<20} {"REF":<15}\n')
            f.write('-' * 100 + '\n')
            
            for record in records:
                if record.user and record.net_salary > 0:
                    f.write(f'{record.bank_account or "":<20} {record.user.name[:30]:<30} ')
                    f.write(f'{float(record.net_salary):>15.2f} {(record.bank_name or "")[:20]:<20} ')
                    f.write(f'PR{record.id:<12}\n')
    
    @staticmethod
    def generate_tax_export(batch_id: int) -> Tuple[bool, Dict]:
        """Generate tax remittance export"""
        result = {
            'success': False,
            'file_path': None,
            'error': None,
            'records': 0,
            'total_tax': Decimal(0)
        }
        
        try:
            batch = PayrollBatch.query.get_or_404(batch_id)
            records = PayrollRecord.query.filter_by(batch_id=batch_id).all()
            
            file_name = f"TAX_REMITTANCE_{batch.payroll_period.replace('-', '')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
            file_path = PayrollExportEngine.EXPORT_BASE_PATH / file_name
            
            PayrollExportEngine.EXPORT_BASE_PATH.mkdir(parents=True, exist_ok=True)
            
            total_tax = Decimal(0)
            with open(file_path, 'w', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=[
                    'Employee Name', 'Employee ID', 'Basic Salary', 'Taxable Income', 
                    'Tax Amount', 'Period', 'Reference'
                ])
                writer.writeheader()
                
                for record in records:
                    if record.tax_deduction > 0:
                        writer.writerow({
                            'Employee Name': record.user.name if record.user else '',
                            'Employee ID': record.user.employee_id if record.user else '',
                            'Basic Salary': float(record.basic_salary),
                            'Taxable Income': float(record.gross_salary),
                            'Tax Amount': float(record.tax_deduction),
                            'Period': record.payroll_period,
                            'Reference': f'PR{record.id}'
                        })
                        total_tax += record.tax_deduction
            
            export = PayrollExport(
                batch_id=batch_id,
                export_type='tax_remittance',
                export_format='csv',
                file_name=file_name,
                file_path=str(file_path),
                file_size=file_path.stat().st_size,
                file_hash=PayrollExportEngine._calculate_file_hash(file_path),
                record_count=len([r for r in records if r.tax_deduction > 0]),
                total_amount=total_tax,
                status='generated'
            )
            db.session.add(export)
            db.session.commit()
            
            result['success'] = True
            result['file_path'] = str(file_path)
            result['records'] = export.record_count
            result['total_tax'] = total_tax
            
            return True, result
            
        except Exception as e:
            logger.error(f"Tax export error: {str(e)}")
            result['error'] = str(e)
            return False, result
    
    @staticmethod
    def generate_pension_export(batch_id: int) -> Tuple[bool, Dict]:
        """Generate pension remittance export"""
        result = {
            'success': False,
            'file_path': None,
            'error': None,
            'total_pension': Decimal(0)
        }
        
        try:
            batch = PayrollBatch.query.get_or_404(batch_id)
            records = PayrollRecord.query.filter_by(batch_id=batch_id).all()
            
            file_name = f"PENSION_REMITTANCE_{batch.payroll_period.replace('-', '')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
            file_path = PayrollExportEngine.EXPORT_BASE_PATH / file_name
            
            PayrollExportEngine.EXPORT_BASE_PATH.mkdir(parents=True, exist_ok=True)
            
            total_pension = Decimal(0)
            with open(file_path, 'w', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=[
                    'Employee Name', 'Pension ID', 'Contribution', 'Period', 'Reference'
                ])
                writer.writeheader()
                
                for record in records:
                    if record.pension_deduction > 0:
                        writer.writerow({
                            'Employee Name': record.user.name if record.user else '',
                            'Pension ID': record.user.employee_id if record.user else '',
                            'Contribution': float(record.pension_deduction),
                            'Period': record.payroll_period,
                            'Reference': f'PR{record.id}'
                        })
                        total_pension += record.pension_deduction
            
            export = PayrollExport(
                batch_id=batch_id,
                export_type='pension',
                export_format='csv',
                file_name=file_name,
                file_path=str(file_path),
                file_size=file_path.stat().st_size,
                file_hash=PayrollExportEngine._calculate_file_hash(file_path),
                record_count=len([r for r in records if r.pension_deduction > 0]),
                total_amount=total_pension,
                status='generated'
            )
            db.session.add(export)
            db.session.commit()
            
            result['success'] = True
            result['file_path'] = str(file_path)
            result['total_pension'] = total_pension
            
            return True, result
            
        except Exception as e:
            logger.error(f"Pension export error: {str(e)}")
            result['error'] = str(e)
            return False, result
    
    @staticmethod
    def _calculate_file_hash(file_path: Path) -> str:
        """Calculate SHA-256 hash of file"""
        sha256_hash = hashlib.sha256()
        with open(file_path, 'rb') as f:
            for byte_block in iter(lambda: f.read(4096), b''):
                sha256_hash.update(byte_block)
        return sha256_hash.hexdigest()
    
    @staticmethod
    def get_export_file(export_id: int) -> Tuple[bool, bytes]:
        """Retrieve generated export file"""
        try:
            export = PayrollExport.query.get_or_404(export_id)
            
            with open(export.file_path, 'rb') as f:
                file_content = f.read()
            
            return True, file_content
            
        except Exception as e:
            logger.error(f"File retrieval error: {str(e)}")
            return False, None
